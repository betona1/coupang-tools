"""쿠팡 로켓그로스 재고추적 서비스

쿠팡 WING Seller Open API 의 재고조회 엔드포인트를 사용해
계정(키)에 등록된 옵션(vendorItemId)의 판매가/재고수량을 조회한다.
원본: \\<DB_HOST>\betona\python\쿠팡그로스재고\cupangpy6.py
"""
import hashlib
import hmac
import json
import re
import ssl
import time
import urllib.error
import urllib.request
from datetime import datetime

from cryptography.fernet import Fernet
from django.conf import settings
from django.db import connections
from django.utils import timezone

from .models import (
    CoupangApiAccount,
    CoupangRocketProduct,
    CoupangInventoryLog,
    CoupangDailySales,
    CoupangRocketConfig,
    CoupangPriceChange,
    CoupangRestock,
    CoupangSettlement,
)

ALLOWED_INTERVALS = (1, 5, 10, 15, 20, 30)
ABNORMAL_DROP = 100   # 단일 측정구간 감소가 이보다 크면 판매 아님(품절/대량조정/입력오류)로 보고 제외
RESTOCK_MIN_DELTA = 3  # 입고로 합산할 최소 증가량 (+1/+2 는 주문취소 반품분으로 보고 제외)
BACKEND_DIR = '/home/joacham/projects/ai100/viewer/gmarket_cpc/backend'
CRON_MARKER = '# COUPANG_ROCKET_STOCK'

API_HOST = "https://api-gateway.coupang.com"
INVENTORY_PATH = "/v2/providers/seller_api/apis/api/v1/marketplace/vendor-items/{vid}/inventories"

UUID_RE = re.compile(r'^[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[1-5][0-9a-fA-F]{3}-[89abAB][0-9a-fA-F]{3}-[0-9a-fA-F]{12}$')
HEX_RE = re.compile(r'^[0-9a-fA-F]+$')
VENDOR_RE = re.compile(r'^[A-Za-z]\d{8,}$')


# ── 전역 설정 (점검 주기) ──
def get_config():
    cfg = CoupangRocketConfig.objects.first()
    if not cfg:
        cfg = CoupangRocketConfig.objects.create(check_interval_min=10)
    return cfg


def set_interval(minutes: int):
    """점검 주기 저장 + crontab 자동 갱신."""
    if minutes not in ALLOWED_INTERVALS:
        raise ValueError(f"허용되지 않는 주기: {minutes} (가능: {ALLOWED_INTERVALS})")
    cfg = get_config()
    cfg.check_interval_min = minutes
    cfg.save(update_fields=['check_interval_min', 'updated_at'])
    _update_cron(minutes)
    return cfg


def _update_cron(minutes: int):
    import subprocess as _sp
    cmd = (f"cd {BACKEND_DIR} && DJANGO_SETTINGS_MODULE=config.settings "
           f"python3 manage.py crawl_coupang_rocket_stock >> {BACKEND_DIR}/coupang_rocket.log 2>&1")
    new_line = f"*/{minutes} * * * * {cmd} {CRON_MARKER}"
    try:
        res = _sp.run(['crontab', '-l'], capture_output=True, text=True)
        existing = res.stdout if res.returncode == 0 else ''
    except Exception:
        existing = ''
    # 기존 쿠팡로켓 라인 제거(마커 + 구버전 명령 둘 다)
    lines = [l for l in existing.split('\n')
             if CRON_MARKER not in l and 'crawl_coupang_rocket_stock' not in l]
    lines = [l for l in lines if l.strip()]
    lines.append(new_line)
    new_crontab = '\n'.join(lines).strip() + '\n'
    _sp.run(['crontab', '-'], input=new_crontab, text=True, capture_output=True)


# ── Fernet 암복호화 (이메일/허브와 동일 키) ──
def _fernet():
    return Fernet(settings.EMAIL_FERNET_KEY.encode())


def encrypt_secret(raw: str) -> str:
    return _fernet().encrypt(raw.encode()).decode()


def decrypt_secret(enc: str) -> str:
    return _fernet().decrypt(enc.encode()).decode()


def clean_token(s) -> str:
    """BOM/제로폭/스마트따옴표/공백 제거"""
    if s is None:
        return ""
    s = str(s).replace('﻿', '')
    s = re.sub(r'[​‌‍⁠﻿]', '', s)
    trans = {0x2018: "'", 0x2019: "'", 0x201C: '"', 0x201D: '"'}
    s = s.translate(trans)
    s = s.strip().strip('"').strip("'")
    return re.sub(r'\s+', '', s)


def validate_account_keys(vendor_id: str, access_key: str, secret_key: str):
    """키 형식 검증 → (ok: bool, errors: list[str]). vendor_id 는 선택(있으면 형식검증)."""
    errors = []
    if vendor_id and not VENDOR_RE.fullmatch(vendor_id):
        errors.append(f"vendorId 형식 오류(예: A00962985): {vendor_id!r}")
    if not UUID_RE.fullmatch(access_key or ''):
        errors.append(f"access-key(UUID) 형식 오류: {access_key!r}")
    if not (secret_key and HEX_RE.fullmatch(secret_key)):
        errors.append("secret-key(hex) 형식 오류")
    return (len(errors) == 0, errors)


# ── 쿠팡 정산(지급내역) ──
SETTLEMENT_PATH = "/v2/providers/marketplace_openapi/apis/api/v1/settlement-histories"


def _signed_get(account, path, query=''):
    """HMAC 서명 GET (query 포함 서명). 반환 (data, error)."""
    secret = decrypt_secret(account.secret_key_enc)
    dt = time.strftime('%y%m%dT%H%M%SZ', time.gmtime())
    msg = f"{dt}GET{path}{query}"
    sig = hmac.new(secret.encode('utf-8'), msg.encode('utf-8'), hashlib.sha256).hexdigest()
    auth = (f"CEA algorithm=HmacSHA256, access-key={account.access_key}, "
            f"signed-date={dt}, signature={sig}")
    url = API_HOST + path + (("?" + query) if query else "")
    req = urllib.request.Request(url)
    req.add_header("Content-Type", "application/json;charset=UTF-8")
    req.add_header("Authorization", auth)
    if account.vendor_id:
        req.add_header("X-Requested-By", account.vendor_id)
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    try:
        with urllib.request.urlopen(req, context=ctx, timeout=30) as resp:
            return json.loads(resp.read().decode(resp.headers.get_content_charset() or 'utf-8')), None
    except urllib.error.HTTPError as e:
        body = ''
        try:
            body = e.read().decode()
        except Exception:
            pass
        return None, f"HTTP {e.code}: {body[:200]}"
    except Exception as e:
        return None, str(e)


def _to_date(s):
    from datetime import datetime
    try:
        return datetime.strptime(s, '%Y-%m-%d').date() if s else None
    except (TypeError, ValueError):
        return None


def sync_settlements(account_id=None, year_month=None):
    """정산 지급내역 동기화. year_month 미지정 시 최근 3개월.
    NDJSON 이벤트 yield."""
    from datetime import date
    accts = CoupangApiAccount.objects.filter(is_active=True)
    if account_id:
        accts = accts.filter(id=account_id)

    if year_month:
        months = [year_month]
    else:
        today = date.today()
        months = []
        y, m = today.year, today.month
        for _ in range(3):
            months.append(f"{y:04d}-{m:02d}")
            m -= 1
            if m == 0:
                m = 12; y -= 1

    total_saved = 0
    for acc in accts:
        for ym in months:
            data, err = _signed_get(acc, SETTLEMENT_PATH, f"revenueRecognitionYearMonth={ym}")
            if err:
                yield {"t": "log", "m": f"❌ {acc.cupang_id} {ym} — {err}"}
                continue
            rows = data if isinstance(data, list) else (data.get('data') or [])
            for s in rows:
                CoupangSettlement.objects.update_or_create(
                    account=acc,
                    settlement_date=_to_date(s.get('settlementDate')),
                    recognition_from=_to_date(s.get('revenueRecognitionDateFrom')),
                    settlement_type=s.get('settlementType', ''),
                    defaults={
                        'revenue_ym': s.get('revenueRecognitionYearMonth', ''),
                        'recognition_to': _to_date(s.get('revenueRecognitionDateTo')),
                        'total_sale': s.get('totalSale') or 0,
                        'service_fee': s.get('serviceFee') or 0,
                        'settlement_target': s.get('settlementTargetAmount') or 0,
                        'settlement_amount': s.get('settlementAmount') or 0,
                        'last_amount': s.get('lastAmount') or 0,
                        'deduction_amount': s.get('deductionAmount') or 0,
                        'seller_service_fee': s.get('sellerServiceFee') or 0,
                        'seller_discount_coupon': s.get('sellerDiscountCoupon') or 0,
                        'downloadable_coupon': s.get('downloadableCoupon') or 0,
                        'store_fee_discount': s.get('storeFeeDiscount') or 0,
                        'debt_of_last_week': s.get('debtOfLastWeek') or 0,
                        'final_amount': s.get('finalAmount') or 0,
                        'bank_name': s.get('bankName') or '',
                        'bank_account': s.get('bankAccount') or '',
                        'bank_holder': s.get('bankAccountHolder') or '',
                        'status': s.get('status') or '',
                        'raw': s,
                    },
                )
                total_saved += 1
            yield {"t": "log", "m": f"✅ {acc.cupang_id} {ym} — {len(rows)}건"}
    yield {"t": "done", "saved": total_saved}


REVENUE_PATH = "/v2/providers/openapi/apis/api/v1/revenue-history"


def get_revenue_history(account, date_from, date_to, max_pages=60):
    """매출내역(주문/상품별) 조회 — vendorId 필수, token 페이지네이션."""
    import urllib.parse
    if not account.vendor_id:
        return {'error': 'vendorId 없음', 'rows': [], 'totals': {}}
    rows = []
    token = ''
    for _ in range(max_pages):
        q = urllib.parse.urlencode({
            'vendorId': account.vendor_id,
            'recognitionDateFrom': date_from,
            'recognitionDateTo': date_to,
            'maxPerPage': 50,
            'token': token,
        })
        data, err = _signed_get(account, REVENUE_PATH, q)
        if err:
            return {'error': err, 'rows': rows, 'totals': _revenue_totals(rows)}
        for r in (data.get('data') or []):
            common = {
                'order_id': r.get('orderId'),
                'sale_type': r.get('saleType'),
                'sale_date': r.get('saleDate'),
                'recognition_date': r.get('recognitionDate'),
                'settlement_date': r.get('settlementDate'),
            }
            for it in (r.get('items') or []):
                rows.append({
                    **common,
                    'product_name': it.get('productName'),
                    'vendor_item_id': it.get('vendorItemId'),
                    'product_id': it.get('productId'),   # 노출상품ID (옵션ID와 달리 반품/로켓/윙 무관 고정)
                    'sku': it.get('externalSellerSkuCode') or '',
                    'quantity': it.get('quantity') or 0,
                    'sale_amount': int(it.get('saleAmount') or 0),
                    'service_fee': it.get('serviceFee') or 0,
                    'service_fee_ratio': it.get('serviceFeeRatio') or 0,
                    'settlement_amount': int(it.get('settlementAmount') or 0),
                })
            # 환불 등 items 없는 배송비 행
            if not (r.get('items')) and r.get('deliveryFee', {}).get('amount'):
                df = r['deliveryFee']
                rows.append({**common, 'product_name': '(배송비)', 'vendor_item_id': None, 'product_id': None, 'sku': '',
                             'quantity': 0, 'sale_amount': int(df.get('amount') or 0),
                             'service_fee': df.get('fee') or 0, 'service_fee_ratio': df.get('feeRatio') or 0,
                             'settlement_amount': int(df.get('settlementAmount') or 0)})
        if not data.get('hasNext'):
            break
        token = data.get('nextToken') or ''
        if not token:
            break
    return {'rows': rows, 'totals': _revenue_totals(rows)}


def _revenue_totals(rows):
    return {
        'count': len(rows),
        'quantity': sum(r['quantity'] for r in rows),
        'sale_amount': sum(r['sale_amount'] for r in rows),
        'service_fee': sum(r['service_fee'] for r in rows),
        'settlement_amount': sum(r['settlement_amount'] for r in rows),
    }


def reconcile_coupang(account, date_from, date_to):
    """쿠팡 매출내역(정산) ↔ order DB(06.쿠팡) 대조.
    매칭키: 쿠팡 orderId = orders_order.bid_number."""
    from collections import defaultdict
    rev = get_revenue_history(account, date_from, date_to)
    if rev.get('error') and not rev['rows']:
        return {'error': rev['error'], 'rows': [], 'totals': {}}

    agg = defaultdict(lambda: {'sale': 0, 'fee': 0, 'settle': 0, 'qty': 0, 'product': '', 'recog': None})
    for r in rev['rows']:
        oid = str(r['order_id'])
        a = agg[oid]
        a['sale'] += r['sale_amount']; a['fee'] += r['service_fee']; a['settle'] += r['settlement_amount']
        a['qty'] += r['quantity']
        if not a['product'] and r['product_name']:
            a['product'] = r['product_name']
        a['recog'] = r['recognition_date']
    ids = list(agg)

    # 반품/취소/환불/교환 상태 판별 (매출대장 부풀림 제외)
    RETURN_COND = ("(order_status LIKE '%%반품%%' OR order_status LIKE '%%취소%%' "
                   "OR order_status LIKE '%%환불%%' OR order_status LIKE '%%교환%%')")
    odb = {}
    if ids:
        with connections['joacham'].cursor() as cur:
            fmt = ','.join(['%s'] * len(ids))
            cur.execute(
                f"""SELECT bid_number,
                          SUM(CASE WHEN {RETURN_COND} THEN 0 ELSE settlement_price END) AS valid_settle,
                          SUM(CASE WHEN {RETURN_COND} THEN settlement_price ELSE 0 END) AS returned_settle,
                          SUM(settlement_price) AS all_settle,
                          MAX(order_status), MAX(order_date)
                    FROM orders_order WHERE site_name='06.쿠팡' AND bid_number IN ({fmt})
                    GROUP BY bid_number""", ids)
            for r in cur.fetchall():
                odb[str(r[0])] = {'settle': int(r[1] or 0), 'returned': int(r[2] or 0),
                                  'all_settle': int(r[3] or 0),
                                  'status': r[4] or '', 'order_date': r[5]}

    rows = []
    for oid in ids:
        a = agg[oid]
        o = odb.get(oid)
        order_settle = o['settle'] if o else None          # 반품/취소 제외 유효정산
        returned = o['returned'] if o else 0
        diff = (a['settle'] - order_settle) if o is not None else None
        if o is None:
            status = 'DB없음'
        elif abs(diff) <= 1:
            status = '일치'
        else:
            status = '불일치'
        rows.append({
            'order_id': oid,
            'product': a['product'],
            'qty': a['qty'],
            'recognition_date': a['recog'],
            'coupang_sale': a['sale'],
            'coupang_fee': a['fee'],
            'coupang_settle': a['settle'],
            'order_settle': order_settle,
            'order_returned': returned,                      # 매출대장에서 제외한 반품/취소액
            'order_status': o['status'] if o else None,
            'diff': diff,
            'status': status,
        })
    # 불일치/DB없음 먼저
    rows.sort(key=lambda r: (0 if r['status'] == '불일치' else 1 if r['status'] == 'DB없음' else 2, r['recognition_date'] or ''))

    matched = [r for r in rows if r['status'] == '일치']
    totals = {
        'count': len(rows),
        'matched': len(matched),
        'mismatch': sum(1 for r in rows if r['status'] == '불일치'),
        'db_missing': sum(1 for r in rows if r['status'] == 'DB없음'),
        'coupang_settle': sum(r['coupang_settle'] for r in rows),
        'order_settle': sum(r['order_settle'] or 0 for r in rows),
        'order_returned': sum(r.get('order_returned') or 0 for r in rows),
        'diff': sum((r['diff'] or 0) for r in rows),
    }
    return {'rows': rows, 'totals': totals}


_IB_STOP = {'세트', '혼합색상', '색상', '사이즈', '무료배송', '특가', '대용량', '국내', '정품',
            '나인조이', 'ero', '개입', '세트입', '묶음', '추가', '옵션', '컬러', '여성', '남성', '아동',
            # 색상
            '브라운', '화이트', '검정', '블랙', '핑크', '그린', '퍼플', '스카이', '레드', '블루',
            '네이비', '그레이', '베이지', '투명', '차콜', '아이보리', '카키', '와인', '옐로우',
            '오렌지', '민트', '라벤더', '실버', '골드', '버건디', '연핑크', '진핑크', '회색', '남색'}


def _ib_tokens(s):
    import re
    return {t.lower() for t in re.findall(r'[가-힣A-Za-z0-9]{2,}', s or '') if t.lower() not in _IB_STOP}


def _load_importbase():
    """importbase 활성 상품 → (customs_name, 토큰셋, 개당원가). 매칭용 1회 로드."""
    try:
        with connections['cost1688'].cursor() as c:
            c.execute("SELECT rate_base FROM exchange_rate WHERE currency='CNY' ORDER BY id DESC LIMIT 1")
            row = c.fetchone(); rate = float(row[0]) if row else 0.0
            c.execute("SELECT cbm_rate_krw, source_cbm_rate_krw FROM sourcing_setting LIMIT 1")
            s = c.fetchone(); mult = (float(s[0]) / float(s[1])) if (s and s[1]) else 1.0
            c.execute("SELECT customs_name, unit_price_cny, customs_fee_unit FROM sourcing_catalog_item WHERE is_active=1")
            items = []
            for nm, cny, cfee in c.fetchall():
                cost = round(float(cny or 0) * rate + float(cfee or 0) * mult)
                items.append((nm or '', _ib_tokens(nm), cost))
            return items
    except Exception:
        return []


def _match_importbase_cost(product_name, items, min_len=4, min_score=4, unique=False):
    """상품명 ↔ 토큰 부분포함 매칭. (name, 개당원가) 또는 None.
    띄어쓰기 차이('팔뚝 압박밴드' vs '팔뚝압박밴드') 대응 위해 토큰 substring 포함 비교.
    min_len/min_score: 사입엑셀(curated)은 3글자까지 허용, importbase(noisy)는 4글자 기본.
    unique=True: 최고점이 약하면(< 4) 동점 후보가 2개 이상일 때 모호하다고 보고 None(오매칭 방지)."""
    ptoks = [t for t in _ib_tokens(product_name) if len(t) >= min_len]
    if not ptoks:
        return None
    best = None; best_score = 0; best_count = 0
    for nm, itoks, cost in items:
        score = 0
        for pt in ptoks:
            for it in itoks:
                if len(it) < min_len:
                    continue
                if pt in it or it in pt:
                    score += min(len(pt), len(it))
        if score > best_score:
            best_score = score; best = (nm, cost); best_count = 1
        elif score == best_score and score > 0:
            best_count += 1
    if best_score < min_score:
        return None
    if unique and best_score < 4 and best_count > 1:   # 약한 점수+동점 → 모호 → 미적용
        return None
    return best


def _match_saip_cost(product_name, saip_items, min_cov=0.6):
    """사입 제품명 기준 커버리지 매칭: 사입 이름의 핵심토큰(3글자+)이 쿠팡 상품명에
    얼마나 포함되는지로 판정. 짧은 이름(모기장)도 정확히 잡고 공통어 오매칭은 배제.
    반환: (사입명, 원가) 또는 None."""
    ptoks = _ib_tokens(product_name)
    if not ptoks:
        return None
    best = None; best_key = (0.0, 0)
    for nm, itoks, cost in saip_items:
        sig = [t for t in itoks if len(t) >= 3]   # 사입명의 의미있는 토큰
        if not sig:
            continue
        matched = 0; mscore = 0
        for st in sig:
            # 매칭쌍 둘 다 3글자+ 일 때만 인정 (수량토큰 '2개'가 '미니화이트보드2개'에 substring 걸리는 오매칭 방지)
            if any((st in pt or pt in st) and min(len(st), len(pt)) >= 3 for pt in ptoks):
                matched += 1; mscore += len(st)
        cov = matched / len(sig)
        if matched >= 1 and cov >= min_cov:
            key = (cov, mscore)
            if key > best_key:
                best_key = key; best = (nm, cost)
    return best


_SAIP_CACHE = {}     # path -> (mtime, [(name, tokens, cost_krw)])
_SAIP_DIR = '/mnt/betona_python/사입상품관리'


def _load_saip_excel():
    """사입상품관리 엑셀(제품명 → 제품원가 KRW=위안×환율×사입수수료) 로드.
    importbase와 동일 포맷 [(name, tokens, cost)] 반환. 파일 mtime 기준 캐시."""
    import os
    import glob
    try:
        files = (glob.glob(os.path.join(_SAIP_DIR, '*사입상품관리*.xlsx'))
                 + glob.glob(os.path.join(_SAIP_DIR, '*소싱관리*.xlsx')))
    except Exception:
        return []
    out = []
    for p in sorted(files):
        if os.path.basename(p).startswith('~$'):
            continue
        try:
            mt = os.path.getmtime(p)
        except OSError:
            continue
        c = _SAIP_CACHE.get(p)
        if c and c[0] == mt:
            out.extend(c[1]); continue
        items = []
        try:
            import openpyxl
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            ws = wb['사입상품기본'] if '사입상품기본' in wb.sheetnames else wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(min_row=1, max_row=3000, values_only=True))
            wb.close()
            hi = next((i for i, r in enumerate(rows)
                       if r and '제품명' in [str(x) for x in r if x is not None]), None)
            if hi is not None:
                hdr = [str(x) if x is not None else '' for x in rows[hi]]
                ni = hdr.index('제품명')
                ci = hdr.index('제품원가') if '제품원가' in hdr else None
                if ci is not None:
                    for r in rows[hi + 1:]:
                        if not r or ni >= len(r) or not r[ni]:
                            continue
                        nm = str(r[ni]).strip()
                        cost = r[ci] if ci < len(r) else None
                        if nm and isinstance(cost, (int, float)) and cost > 0:
                            items.append((nm, _ib_tokens(nm), round(float(cost))))
        except Exception:
            items = []
        _SAIP_CACHE[p] = (mt, items)
        out.extend(items)
    return out


def get_coupang_product_settlement(seller_alias=None, date_from=None, date_to=None):
    """order DB(06.쿠팡)에서 상품별 정산 집계 — 사업자/기간 필터, 반품·취소 제외, 원가·이익 포함."""
    RET = ("(order_status LIKE '%%반품%%' OR order_status LIKE '%%취소%%' "
           "OR order_status LIKE '%%환불%%' OR order_status LIKE '%%교환%%')")
    where = ["site_name='06.쿠팡'"]
    params = []
    if seller_alias:
        where.append("seller_alias=%s"); params.append(seller_alias)
    if date_from:
        where.append("order_date>=%s"); params.append(date_from)
    if date_to:
        where.append("order_date<=%s"); params.append(date_to)
    wsql = ' AND '.join(where)
    DEFAULT_SHIP = 2620   # 기본배송비(부가세 포함) — real_shipping_fee 없을 때 적용
    sql = f"""
        SELECT COALESCE(NULLIF(product_code,''), product_name) AS pkey,
               MAX(product_name) AS pname, MAX(seller_alias) AS alias,
               MAX(product_seller_code) AS scode, MAX(product_code) AS pcode,
               SUM(CASE WHEN NOT {RET} THEN quantity ELSE 0 END) AS qty,
               SUM(CASE WHEN NOT {RET} THEN settlement_price ELSE 0 END) AS settle,
               SUM(CASE WHEN NOT {RET} THEN payment_price ELSE 0 END) AS pay,
               SUM(CASE WHEN NOT {RET} THEN supply_price ELSE 0 END) AS supply,
               SUM(CASE WHEN {RET} THEN 1 ELSE 0 END) AS cancel_cnt,
               SUM(CASE WHEN {RET} THEN settlement_price ELSE 0 END) AS cancel_amt,
               COUNT(*) AS cnt,
               SUM(CASE WHEN NOT {RET}
                        THEN (CASE WHEN COALESCE(real_shipping_fee,0)>0 THEN real_shipping_fee ELSE {DEFAULT_SHIP} END)
                        ELSE 0 END) AS shipping
        FROM orders_order WHERE {wsql}
        GROUP BY pkey HAVING qty > 0 OR cancel_cnt > 0
        ORDER BY settle DESC
    """
    raw = []
    with connections['joacham'].cursor() as cur:
        cur.execute(sql, params)
        raw = cur.fetchall()
    # 인덱스: 0=pkey 1=pname 2=alias 3=scode 4=pcode 5=qty 6=settle 7=pay 8=supply 9=cancel_cnt 10=cancel_amt 11=cnt 12=shipping
    # 원가 매핑 테이블 로드 (노출상품ID / S코드 기준)
    from .models import CoupangProductCostMap
    cmap_exp, cmap_code = {}, {}
    for m in CoupangProductCostMap.objects.all().values('exposure_id', 'product_seller_code', 'unit_cost', 'bundle_size', 'ship_excluded', 'importbase_name'):
        if m['exposure_id']:
            cmap_exp[m['exposure_id']] = m
        if m['product_seller_code']:
            cmap_code[m['product_seller_code']] = m
    # importbase + 사입상품관리 엑셀 로드 (원가 0이거나 매핑 제안용) — 제품명 토큰 매칭
    need_match = any(int(r[8] or 0) == 0 and int(r[5] or 0) > 0 for r in raw)
    saip_items = _load_saip_excel() if need_match else []
    ib_only = _load_importbase() if need_match else []
    # 광고비 — 노출상품ID 기준 집계 (옵션ID/로켓ID 무관). 정산 상품명 인덱스로 상품명 보조매칭 지원.
    name2exp = []
    _seen_exp = set()
    for r in raw:
        pc = str(r[4] or '').strip()
        if pc and pc not in _seen_exp:
            _seen_exp.add(pc)
            name2exp.append((pc, {t for t in _ib_tokens(r[1] or '') if len(t) >= 3}))
    ad_by_exp = get_ad_cost_by_exposure(date_from, date_to, name2exp=name2exp) if (date_from and date_to) else {}

    rows = []
    tot = {'qty': 0, 'settle': 0, 'pay': 0, 'supply': 0, 'shipping': 0, 'profit': 0,
           'ad_cost': 0, 'ad_sales': 0, 'profit_ad': 0, 'cancel_cnt': 0, 'cancel_amt': 0}
    for r in raw:
        pname = r[1]; scode = r[3] or ''; pcode = r[4] or ''
        qty = int(r[5] or 0); settle = int(r[6] or 0); supply = int(r[8] or 0)
        cost_src = 'order'; ib_name = ''; unit_cost = 0; bundle = 1
        ib_sug_cost = 0; ib_sug_name = ''; ib_sug_src = ''
        # 제품명 토큰 매칭 제안 — 사입엑셀 우선, 없으면 importbase
        sm = _match_saip_cost(pname, saip_items) if saip_items else None
        if sm:
            ib_sug_src = '사입'
        elif ib_only:
            sm = _match_importbase_cost(pname, ib_only)
            if sm:
                ib_sug_src = '1688'
        if sm:
            ib_sug_cost, ib_sug_name = sm[1], sm[0]
        cm = cmap_exp.get(pcode) or cmap_code.get(scode)
        if cm and cm['unit_cost']:                       # 1순위: 원가 매핑 테이블
            unit_cost = cm['unit_cost']; bundle = cm['bundle_size'] or 1
            supply = unit_cost * bundle * qty
            cost_src = 'map'; ib_name = cm.get('importbase_name') or ''
        elif supply > 0:                                  # 2순위: 주문DB 공급가
            cost_src = 'order'
        elif qty > 0 and ib_sug_cost:                     # 3순위: 사입엑셀/importbase 자동매칭(판매단위당 원가)
            unit_cost = ib_sug_cost; bundle = 1
            supply = ib_sug_cost * qty
            cost_src = 'saip' if ib_sug_src == '사입' else 'importbase'; ib_name = ib_sug_name
        else:
            cost_src = 'none'
        shipping = int(r[12] or 0)                 # 배송비(real_shipping_fee, 없으면 건당 기본 2620)
        ship_excl = bool(cm and cm.get('ship_excluded'))
        if ship_excl:                              # 오너클랜배송 등 우리가 안 보내는 상품 → 배송비 0
            shipping = 0
        profit = settle - supply - shipping        # 이익 = 정산 − 원가 − 배송비
        # 광고비 (노출상품ID 매칭)
        ad = ad_by_exp.get(pcode) if pcode else None
        ad_cost = ad['ad_cost'] if ad else 0
        ad_sales = ad['ad_sales'] if ad else 0
        roas = round(ad_sales / ad_cost * 100) if ad_cost else 0      # 광고수익률(%) = 광고매출/광고비
        acos = round(ad_cost / settle * 100, 1) if settle else 0      # 광고비/매출 비중(%)
        profit_ad = profit - ad_cost                                  # 광고후이익 = 이익 − 광고비
        row = {
            'product_name': pname, 'seller_alias': r[2], 'product_seller_code': scode, 'exposure_id': pcode,
            'qty': qty, 'settle': settle, 'pay': int(r[7] or 0), 'supply': supply, 'shipping': shipping,
            'unit_cost': unit_cost, 'bundle_size': bundle, 'ship_excluded': ship_excl,
            'profit': profit, 'margin': round(profit / settle * 100, 1) if settle else 0,
            'ad_cost': ad_cost, 'ad_sales': ad_sales, 'roas': roas, 'acos': acos, 'profit_ad': profit_ad,
            'cost_source': cost_src, 'import_name': ib_name,
            'import_suggest_cost': ib_sug_cost, 'import_suggest_name': ib_sug_name, 'import_suggest_src': ib_sug_src,
            'cancel_cnt': int(r[9] or 0), 'cancel_amt': int(r[10] or 0), 'cnt': int(r[11] or 0),
        }
        rows.append(row)
        for k in ('qty', 'settle', 'pay', 'supply', 'shipping', 'ad_cost', 'ad_sales', 'cancel_cnt', 'cancel_amt'):
            tot[k] += row[k]
        tot['profit'] += profit
        tot['profit_ad'] += profit_ad
    rows.sort(key=lambda x: x['settle'], reverse=True)
    tot['margin'] = round(tot['profit'] / tot['settle'] * 100, 1) if tot['settle'] else 0
    tot['roas'] = round(tot['ad_sales'] / tot['ad_cost'] * 100) if tot['ad_cost'] else 0
    return {'rows': rows, 'totals': tot}


# ── 상품 원가/번들 매핑 CRUD ──
def get_product_cost_map():
    from .models import CoupangProductCostMap
    return list(CoupangProductCostMap.objects.all().order_by('-updated_at').values(
        'id', 'exposure_id', 'product_seller_code', 'product_name', 'unit_cost', 'bundle_size', 'ship_excluded', 'importbase_name', 'memo'))


def upsert_product_cost_map(exposure_id='', product_seller_code='', product_name='', unit_cost=0, bundle_size=1, ship_excluded=None, importbase_name='', memo=''):
    """노출상품ID 우선, 없으면 S코드 기준 upsert. ship_excluded=None이면 기존값 유지."""
    from .models import CoupangProductCostMap
    exposure_id = (exposure_id or '').strip(); product_seller_code = (product_seller_code or '').strip()
    if not exposure_id and not product_seller_code:
        raise ValueError('노출상품ID 또는 S코드 필요')
    lookup = {'exposure_id': exposure_id} if exposure_id else {'product_seller_code': product_seller_code}
    defaults = {
        'exposure_id': exposure_id, 'product_seller_code': product_seller_code,
        'product_name': product_name, 'unit_cost': int(unit_cost or 0),
        'bundle_size': max(1, int(bundle_size or 1)), 'importbase_name': importbase_name, 'memo': memo,
    }
    if ship_excluded is not None:
        defaults['ship_excluded'] = bool(ship_excluded)
    obj, _ = CoupangProductCostMap.objects.update_or_create(defaults=defaults, **lookup)
    return obj.id


def delete_product_cost_map(map_id):
    from .models import CoupangProductCostMap
    CoupangProductCostMap.objects.filter(id=map_id).delete()


# ── 쿠팡 광고비 (WING 광고리포트 엑셀 업로드 + 노출상품ID 집계) ──
# 탐지 우선순위 순서(앞에서 배정된 컬럼은 뒤에서 재사용 안 함). 키워드는 구체적으로.
_AD_COLS = [
    ('exposure_id', ['노출상품id', '노출상품번호', '노출상품아이디', '노출상품코드', '광고집행상품id', '노출상품', 'productid']),
    ('vendor_item_id', ['옵션id', 'vendoritemid', '옵션아이디', '광고옵션id', '옵션번호', '아이템id']),
    ('ad_date', ['날짜', '일자', '광고일', '집행일', '노출일', 'date']),
    ('campaign_name', ['캠페인명', '캠페인', 'campaign']),
    ('ad_type', ['광고유형', '광고타입', 'adtype']),
    ('product_name', ['광고상품명', '상품명', 'productname']),
    ('ad_cost', ['광고비', '집행광고비', '광고비용', '광고비집행', 'adcost', 'spend']),
    ('ad_sales', ['총전환매출액', '전환매출액', '총전환매출']),   # '광고전환매출발생 상품명'(텍스트) 오매칭 방지
    ('ad_orders', ['총판매수량', '총전환판매수', '광고전환판매수', '전환판매수', '판매수량', '총주문수']),
    ('impressions', ['노출수', 'impressions', 'impression']),
    ('clicks', ['클릭수', 'clicks', 'click']),
]
_AD_INT_MAX = 2147483647


def _ad_norm(s):
    import re
    return re.sub(r'[\s_()/\-.,]', '', str(s if s is not None else '')).lower()


def _ad_to_int(v):
    try:
        if v in (None, ''):
            return 0
        n = int(round(float(str(v).replace(',', '').replace('₩', '').replace('%', '').strip())))
        return max(0, min(n, _AD_INT_MAX))   # int32 범위 클램프(컬럼 오매핑 시 DB 에러 방지)
    except (ValueError, TypeError):
        return 0


def _ad_parse_date(v, default):
    from datetime import datetime, date as _date
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, _date):
        return v
    s = str(v or '').strip()
    if s.endswith('.0'):          # 엑셀 숫자셀 '20260612.0'
        s = s[:-2]
    digits = ''.join(ch for ch in s if ch.isdigit())
    if len(digits) == 8:          # YYYYMMDD
        try:
            return datetime.strptime(digits, '%Y%m%d').date()
        except ValueError:
            pass
    s2 = s.replace('.', '-').replace('/', '-')[:10]
    try:
        return datetime.strptime(s2, '%Y-%m-%d').date()
    except ValueError:
        return default


def upload_coupang_ad_excel(file_obj, cupang_id, default_date=None):
    """WING 광고관리 리포트(xlsx/csv) → CoupangAdCost. 헤더 키워드 자동탐지.
    옵션ID·노출상품ID 둘 다 저장(집계는 노출상품ID 기준). 같은 cupang_id+날짜는 재업로드 시 교체."""
    import openpyxl
    from datetime import date as _date
    from .models import CoupangAdCost
    if default_date is None:
        default_date = _date.today()
    elif isinstance(default_date, str):
        default_date = _ad_parse_date(default_date, _date.today())

    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return {'error': '빈 파일', 'inserted': 0}

    # 헤더행 탐색(광고비 + 노출/클릭 포함)
    hi = None
    for i, r in enumerate(rows[:40]):
        cells = [_ad_norm(c) for c in (r or [])]
        if any(('광고비' in c or 'adcost' in c or 'spend' in c) for c in cells) and \
           any(('노출' in c or '클릭' in c or 'click' in c or 'impress' in c) for c in cells):
            hi = i
            break
    if hi is None:
        return {'error': '헤더 인식 실패 — 광고비/노출수 컬럼이 있는 WING 광고리포트를 올려주세요', 'inserted': 0}

    hdr = [_ad_norm(c) for c in rows[hi]]
    used = set()

    def find_col(keys):
        for k in keys:                       # 키워드 우선순위대로(구체적 키워드 먼저)
            for j, h in enumerate(hdr):
                if j in used:
                    continue
                if k in h:
                    used.add(j)
                    return j
        return None
    col = {f: find_col(ks) for f, ks in _AD_COLS}
    if col['ad_cost'] is None:
        return {'error': '광고비 컬럼 못 찾음', 'inserted': 0}

    def g(r, f):
        j = col.get(f)
        return r[j] if (j is not None and j < len(r) and r[j] is not None) else ''

    objs = []
    dates = set()
    for r in rows[hi + 1:]:
        if not r:
            continue
        exp = str(g(r, 'exposure_id') or '').strip().split('.')[0]
        vid = str(g(r, 'vendor_item_id') or '').strip().split('.')[0]
        if not exp and not vid:
            continue
        d = _ad_parse_date(g(r, 'ad_date'), default_date) if col['ad_date'] is not None else default_date
        objs.append(CoupangAdCost(
            cupang_id=cupang_id, ad_date=d, exposure_id=exp, vendor_item_id=vid,
            campaign_name=str(g(r, 'campaign_name') or '')[:200], ad_type=str(g(r, 'ad_type') or '')[:30],
            product_name=str(g(r, 'product_name') or '')[:300],
            impressions=_ad_to_int(g(r, 'impressions')), clicks=_ad_to_int(g(r, 'clicks')),
            ad_cost=_ad_to_int(g(r, 'ad_cost')), ad_orders=_ad_to_int(g(r, 'ad_orders')),
            ad_sales=_ad_to_int(g(r, 'ad_sales')), source='excel'))
        dates.add(d)
    if objs:
        CoupangAdCost.objects.filter(cupang_id=cupang_id, ad_date__in=list(dates)).delete()
        CoupangAdCost.objects.bulk_create(objs, batch_size=500)
    detected = {f: (rows[hi][j] if j is not None else None) for f, j in col.items()}
    return {'inserted': len(objs), 'dates': sorted(str(x) for x in dates), 'detected': detected,
            'total_ad_cost': sum(o.ad_cost for o in objs)}


def _build_vid2exp():
    """옵션ID→노출상품ID 통합 맵. 우선순위: 카탈로그(CoupangVidMap) > 재고추적(CoupangRocketProduct)."""
    from .models import CoupangVidMap, CoupangRocketProduct
    vid2exp = {}
    # 재고추적(보조)
    for p in CoupangRocketProduct.objects.exclude(seller_product_id='').values(
            'vendor_item_id', 'marketplace_vendor_item_id', 'seller_product_id'):
        sp = str(p['seller_product_id'] or '').strip()
        if not sp:
            continue
        if p['vendor_item_id']:
            vid2exp[str(p['vendor_item_id']).strip()] = sp
        if p.get('marketplace_vendor_item_id'):
            vid2exp[str(p['marketplace_vendor_item_id']).strip()] = sp
    # 카탈로그(우선) — 전 상품 순회 캐시
    for v, e in CoupangVidMap.objects.values_list('vendor_item_id', 'exposure_id'):
        if v and e:
            vid2exp[str(v).strip()] = str(e).strip()
    return vid2exp


def sync_vid_exposure_map(account, max_pages=600, delay=0.1):
    """전 상품(seller-products) 목록+상세 순회 → 옵션ID→노출상품ID 카탈로그(CoupangVidMap) 구축.
    윙(mp_vid)/로켓(rg_vid) 옵션ID 모두 같은 노출ID(productId)로 매핑. NDJSON 이벤트 yield."""
    import time as _t
    from .models import CoupangVidMap
    if not (account.access_key and account.secret_key_enc):
        yield {'t': 'log', 'm': '❌ Open API 키 없음'}
        yield {'t': 'done', 'mapped': 0, 'products': 0}
        return
    pairs = []
    token = ''
    page = 0
    while page < max_pages:
        page += 1
        q = f"vendorId={account.vendor_id}&maxPerPage=50" + (f"&nextToken={token}" if token else "")
        d, err = _signed_get(account, SELLER_PRODUCTS_PATH, q)
        if err:
            yield {'t': 'log', 'm': f'목록 {page}p 실패: {err}'}
            break
        for it in (d.get('data') or []):
            pid = str(it.get('productId') or '').strip()
            spid = str(it.get('sellerProductId') or '').strip()
            if pid and spid:
                pairs.append((pid, spid))
        if page % 5 == 0:
            yield {'t': 'log', 'm': f'상품목록 {page}페이지 / 누적 {len(pairs)}개'}
        token = d.get('nextToken')
        if not token:
            break
    yield {'t': 'log', 'm': f'상품 {len(pairs)}개 — 옵션ID 조회 시작...'}
    mapped = 0
    for i, (pid, spid) in enumerate(pairs, 1):
        d, err = _signed_get(account, f"{SELLER_PRODUCTS_PATH}/{spid}")
        if not err and d:
            for it in ((d.get('data') or {}).get('items') or []):
                for vid in (str((it.get('marketplaceItemData') or {}).get('vendorItemId') or '').strip(),
                            str((it.get('rocketGrowthItemData') or {}).get('vendorItemId') or '').strip()):
                    if vid and vid != 'None':
                        CoupangVidMap.objects.update_or_create(
                            vendor_item_id=vid,
                            defaults={'exposure_id': pid, 'seller_product_id': spid, 'cupang_id': account.cupang_id})
                        mapped += 1
        if i % 50 == 0:
            yield {'t': 'log', 'm': f'{i}/{len(pairs)} 상품 처리 · 옵션 {mapped}개 매핑'}
        _t.sleep(delay)
    yield {'t': 'log', 'm': f'✅ 완료 — 상품 {len(pairs)} / 옵션ID {mapped}개 매핑'}
    yield {'t': 'done', 'products': len(pairs), 'mapped': mapped}


def _match_name_exposure(name, name2exp, min_score=6):
    """광고 상품명 → 정산 상품명 토큰매칭으로 노출상품ID 추정(보수적).
    name2exp: [(exposure_id, {정산명 3글자+ 토큰})]. 최고점>=min_score & 단독1위일 때만."""
    toks = [t for t in _ib_tokens(name) if len(t) >= 3]
    if not toks:
        return None
    best = None; bs = 0; bc = 0
    for exp, etoks in name2exp:
        sc = 0
        for pt in toks:
            for et in etoks:
                if pt in et or et in pt:
                    sc += min(len(pt), len(et))
                    break
        if sc > bs:
            bs = sc; best = exp; bc = 1
        elif sc == bs and sc > 0:
            bc += 1
    return best if (bs >= min_score and bc == 1) else None


def get_ad_cost_by_exposure(date_from, date_to, cupang_id=None, name2exp=None):
    """노출상품ID별 광고비/광고매출 집계. 광고보고서엔 옵션ID만 있으므로
    옵션ID→노출상품ID(카탈로그 CoupangVidMap + 재고추적)로 환산해 합산.
    환산 불가 시 name2exp(정산 상품명 인덱스)로 상품명 보조매칭, 그래도 안되면 'VID:..' 키."""
    from collections import defaultdict
    from .models import CoupangAdCost
    vid2exp = _build_vid2exp()
    name_cache = {}
    qs = CoupangAdCost.objects.filter(ad_date__gte=date_from, ad_date__lte=date_to)
    if cupang_id:
        qs = qs.filter(cupang_id=cupang_id)
    agg = defaultdict(lambda: {'ad_cost': 0, 'ad_sales': 0, 'impressions': 0, 'clicks': 0, 'ad_orders': 0, 'vids': set(), 'named': 0})
    for r in qs.values('exposure_id', 'vendor_item_id', 'product_name', 'ad_cost', 'ad_sales', 'impressions', 'clicks', 'ad_orders'):
        vid = (r['vendor_item_id'] or '').strip()
        exp = (r['exposure_id'] or '').strip() or vid2exp.get(vid, '')
        named = False
        if not exp and name2exp and r['product_name']:    # 상품명 보조매칭
            nm = r['product_name']
            if nm not in name_cache:
                name_cache[nm] = _match_name_exposure(nm, name2exp) or ''
            exp = name_cache[nm]
            named = bool(exp)
        key = exp or ('VID:' + vid)
        a = agg[key]
        for k in ('ad_cost', 'ad_sales', 'impressions', 'clicks', 'ad_orders'):
            a[k] += r[k] or 0
        if vid:
            a['vids'].add(vid)
        if named:
            a['named'] += r['ad_cost'] or 0
    return agg


def get_ad_efficiency(date_from, date_to, cupang_id=None):
    """광고비효율 대시보드 집계 — CoupangAdCost 기준.
    요약(광고비/광고매출/ROAS/CTR/CPC) + 계정별 + 상품별(노출ID 또는 상품명) + 일별 추이."""
    from collections import defaultdict
    from .models import CoupangAdCost

    qs = CoupangAdCost.objects.filter(ad_date__gte=date_from, ad_date__lte=date_to)
    if cupang_id:
        qs = qs.filter(cupang_id=cupang_id)

    def _eff(cost, sales, impr, clk):
        return {
            'ad_cost': cost, 'ad_sales': sales, 'impressions': impr, 'clicks': clk,
            'roas': round(sales / cost * 100) if cost else 0,          # 광고수익률
            'acos': round(cost / sales * 100, 1) if sales else 0,      # 광고비/매출
            'ctr': round(clk / impr * 100, 2) if impr else 0,          # 클릭률
            'cpc': round(cost / clk) if clk else 0,                    # 클릭당 광고비
        }

    tot = defaultdict(int)
    by_acc = defaultdict(lambda: defaultdict(int))
    by_prod = defaultdict(lambda: {'name': '', 'exposure_id': '', 'ad_cost': 0, 'ad_sales': 0,
                                   'impressions': 0, 'clicks': 0, 'ad_orders': 0})
    by_day = defaultdict(lambda: defaultdict(int))

    for r in qs.values('cupang_id', 'ad_date', 'exposure_id', 'product_name',
                       'ad_cost', 'ad_sales', 'impressions', 'clicks', 'ad_orders'):
        c, s, im, ck, od = (r['ad_cost'] or 0, r['ad_sales'] or 0, r['impressions'] or 0,
                            r['clicks'] or 0, r['ad_orders'] or 0)
        for k, v in (('ad_cost', c), ('ad_sales', s), ('impressions', im), ('clicks', ck), ('ad_orders', od)):
            tot[k] += v
            by_acc[r['cupang_id']][k] += v
        # 상품 키: 노출상품ID 우선, 없으면 상품명(옵션접미사 제거)
        exp = (r['exposure_id'] or '').strip()
        base = (r['product_name'] or '').split(',')[0].strip()
        gkey = exp or base or '(미상)'
        g = by_prod[gkey]
        g['name'] = g['name'] or base
        g['exposure_id'] = g['exposure_id'] or exp
        for k, v in (('ad_cost', c), ('ad_sales', s), ('impressions', im), ('clicks', ck), ('ad_orders', od)):
            g[k] += v
        ds = str(r['ad_date'])
        by_day[ds]['ad_cost'] += c
        by_day[ds]['ad_sales'] += s

    accounts = sorted(
        [{'cupang_id': cid, **_eff(d['ad_cost'], d['ad_sales'], d['impressions'], d['clicks']),
          'ad_orders': d['ad_orders']} for cid, d in by_acc.items()],
        key=lambda x: -x['ad_cost'])
    products = sorted(
        [{'name': g['name'], 'exposure_id': g['exposure_id'], 'ad_orders': g['ad_orders'],
          **_eff(g['ad_cost'], g['ad_sales'], g['impressions'], g['clicks'])}
         for g in by_prod.values() if g['ad_cost'] > 0],
        key=lambda x: -x['ad_cost'])
    daily = [{'date': d[5:], 'full_date': d, 'ad_cost': by_day[d]['ad_cost'], 'ad_sales': by_day[d]['ad_sales']}
             for d in sorted(by_day)]
    # 저효율(ROAS<100) 상품 광고비 합
    low_roas_cost = sum(p['ad_cost'] for p in products if p['roas'] < 100)
    return {
        'range': {'from': str(date_from), 'to': str(date_to)},
        'totals': {**_eff(tot['ad_cost'], tot['ad_sales'], tot['impressions'], tot['clicks']),
                   'ad_orders': tot['ad_orders'], 'product_count': len(products),
                   'low_roas_cost': low_roas_cost},
        'accounts': accounts,
        'products': products,
        'daily': daily,
    }


def get_product_reviews(product_key, limit=200):
    """노출상품ID의 쿠팡 리뷰. **내용 있는 리뷰만**(별점만 있는 건 제외) + 최근순.
    평균별점/분포는 별점 있는 전체 기준(요약), 목록은 내용 있는 것만."""
    dist = {5: 0, 4: 0, 3: 0, 2: 0, 1: 0}
    rated_total = 0
    rsum = 0.0
    rows = []
    try:
        with connections['joacham'].cursor() as c:
            # 별점 요약: 별점 있는 전체
            c.execute("SELECT rating FROM coupang_review WHERE product_id=%s AND rating IS NOT NULL",
                      [str(product_key)])
            for (rating,) in c.fetchall():
                rated_total += 1
                rsum += float(rating)
                b = int(round(float(rating)))
                if b in dist:
                    dist[b] += 1
            # 목록: 내용 있는 리뷰만, 최근순
            c.execute(
                "SELECT rating, headline, content, reviewer, review_date, helpful_count "
                "FROM coupang_review WHERE product_id=%s AND content<>'' AND content IS NOT NULL "
                "ORDER BY review_date DESC, id DESC LIMIT %s",
                [str(product_key), limit])
            for rating, headline, content, reviewer, rdate, helpful in c.fetchall():
                rows.append({
                    'rating': float(rating) if rating is not None else None,
                    'headline': headline or '', 'content': content or '',
                    'reviewer': reviewer or '', 'review_date': rdate or '',
                    'helpful_count': helpful or 0,
                })
    except Exception as e:
        return {'product_key': product_key, 'count': 0, 'avg': 0, 'dist': dist,
                'reviews': [], 'error': str(e)}
    return {
        'product_key': product_key,
        'count': len(rows),                 # 내용 있는 리뷰 수
        'rated_count': rated_total,          # 별점 매긴 전체 수(요약 기준)
        'avg': round(rsum / rated_total, 2) if rated_total else 0,
        'dist': dist, 'reviews': rows,
    }


def get_review_report():
    """리뷰 증가 리포트 — 직전 스냅샷 대비 증가분(상품별). 일일 크롤 후 팝업용.
    스냅샷(coupang_review_snapshot)에서 각 상품의 최신 vs 직전 cnt 차이."""
    rows = []
    latest_date = None
    try:
        with connections['joacham'].cursor() as c:
            # 상품별 최신 2개 스냅샷
            c.execute("SELECT product_id, snapshot_date, cnt FROM coupang_review_snapshot "
                      "ORDER BY product_id, snapshot_date DESC")
            from collections import defaultdict
            snaps = defaultdict(list)
            for pid, d, cnt in c.fetchall():
                snaps[str(pid)].append((d, int(cnt)))
        # 증가분 계산
        prod_rows = []
        for pid, lst in snaps.items():
            cur_d, cur = lst[0]
            prev = lst[1][1] if len(lst) > 1 else cur   # 직전 없으면(첫 스냅샷=baseline) 증가0 처리
            inc = cur - prev
            if latest_date is None or str(cur_d) > str(latest_date):
                latest_date = cur_d
            if inc > 0:
                prod_rows.append({'product_id': pid, 'total': cur, 'increase': inc})
        # 상품명
        if prod_rows:
            pids = [r['product_id'] for r in prod_rows]
            ph = ','.join(['%s'] * len(pids))
            with connections['default'].cursor() as c2:
                c2.execute(f"SELECT seller_product_id, MIN(product_name) FROM cupang_rocket_product "
                           f"WHERE seller_product_id IN ({ph}) GROUP BY seller_product_id", pids)
                names = {str(a): b for a, b in c2.fetchall()}
            for r in prod_rows:
                r['product_name'] = names.get(r['product_id'], r['product_id'])
        prod_rows.sort(key=lambda x: -x['increase'])
        rows = prod_rows
    except Exception:
        pass
    return {'date': str(latest_date) if latest_date else '',
            'total_increase': sum(r['increase'] for r in rows), 'products': rows}


def get_reviews_summary(product_keys):
    """여러 노출ID의 리뷰 요약(건수+평균) — 목록 뱃지용."""
    out = {}
    if not product_keys:
        return out
    try:
        ph = ','.join(['%s'] * len(product_keys))
        with connections['joacham'].cursor() as c:
            c.execute(f"SELECT product_id, COUNT(*), ROUND(AVG(rating),1) "
                      f"FROM coupang_review WHERE product_id IN ({ph}) GROUP BY product_id",
                      [str(k) for k in product_keys])
            for pid, cnt, avg in c.fetchall():
                out[str(pid)] = {'count': cnt, 'avg': float(avg) if avg else 0}
    except Exception:
        pass
    return out


def get_product_daily_sales(product_key, account_id=None, days=30):
    """베스트 상품(노출상품ID 묶음)의 최근 N일 일별 판매량 + 주말/평일 비교.
    product_key: seller_product_id 또는 '_<CoupangRocketProduct.id>'(노출ID 없는 단일옵션)."""
    from datetime import date, timedelta
    from .models import CoupangRocketProduct, CoupangDailySales

    # product_key → vendor_item_id 목록 해석
    prods = CoupangRocketProduct.objects.all()
    if account_id:
        prods = prods.filter(account_id=account_id)
    if product_key.startswith('_'):
        prods = prods.filter(id=product_key[1:])
    else:
        prods = prods.filter(seller_product_id=product_key)
    vids = list(prods.values_list('vendor_item_id', flat=True))
    pname = (prods.first().product_name if prods.first() else '') or ''
    if not vids:
        return {'product_key': product_key, 'product_name': pname, 'days': [], 'summary': {}}

    end = date.today()
    start = end - timedelta(days=days - 1)
    # vid별 일별 판매 합산
    by_date = {}
    for r in CoupangDailySales.objects.filter(vendor_item_id__in=vids, date__gte=start, date__lte=end).values('date', 'sold_quantity'):
        by_date[r['date']] = by_date.get(r['date'], 0) + (r['sold_quantity'] or 0)

    WD = ['월', '화', '수', '목', '금', '토', '일']
    out = []
    wk_qty = wk_days = wd_qty = wd_days = 0
    for i in range(days):
        d = start + timedelta(days=i)
        wd = d.weekday()                 # 0=월 .. 5=토,6=일
        is_weekend = wd >= 5
        q = by_date.get(d, 0)
        out.append({'date': d.strftime('%m/%d'), 'full_date': str(d), 'weekday': WD[wd],
                    'is_weekend': is_weekend, 'qty': q})
        if is_weekend:
            wk_qty += q; wk_days += 1
        else:
            wd_qty += q; wd_days += 1

    total = wk_qty + wd_qty
    weekend_avg = round(wk_qty / wk_days, 1) if wk_days else 0
    weekday_avg = round(wd_qty / wd_days, 1) if wd_days else 0
    all_avg = round(total / days, 1) if days else 0
    # 판정: 주말 일평균이 평일 일평균보다 20%+ 높으면 주말형, 반대면 평일형
    if weekday_avg and weekend_avg >= weekday_avg * 1.2:
        verdict, ratio = '주말형', round(weekend_avg / weekday_avg, 2) if weekday_avg else 0
    elif weekend_avg and weekday_avg >= weekend_avg * 1.2:
        verdict, ratio = '평일형', round(weekday_avg / weekend_avg, 2) if weekend_avg else 0
    else:
        verdict, ratio = '고른편', 1.0
    return {
        'product_key': product_key, 'product_name': pname, 'option_count': len(vids),
        'days': out,
        'summary': {
            'total': total, 'weekend_qty': wk_qty, 'weekday_qty': wd_qty,
            'weekend_days': wk_days, 'weekday_days': wd_days,
            'weekend_avg': weekend_avg, 'weekday_avg': weekday_avg, 'all_avg': all_avg,
            'verdict': verdict, 'ratio': ratio,
        },
    }


def get_ad_campaigns(date_from, date_to, cupang_id=None):
    """캠페인별 광고 집계 — CoupangAdCost 기준. 캠페인명으로 묶어 광고비/ROAS + 집행상품 목록.
    각 캠페인에 해당 계정의 설정 변경이력(change_history)도 첨부."""
    from collections import defaultdict
    from .models import CoupangAdCost, CoupangAdChange

    qs = CoupangAdCost.objects.filter(ad_date__gte=date_from, ad_date__lte=date_to)
    if cupang_id:
        qs = qs.filter(cupang_id=cupang_id)

    camps = defaultdict(lambda: {'ad_cost': 0, 'ad_sales': 0, 'impressions': 0, 'clicks': 0,
                                 'ad_orders': 0, 'products': defaultdict(lambda: {'name': '', 'exposure_id': '',
                                 'ad_cost': 0, 'ad_sales': 0, 'clicks': 0})})
    for r in qs.values('cupang_id', 'campaign_name', 'exposure_id', 'product_name',
                       'ad_cost', 'ad_sales', 'impressions', 'clicks', 'ad_orders'):
        cname = (r['campaign_name'] or '(캠페인 미지정)').strip()
        c = camps[cname]
        for k in ('ad_cost', 'ad_sales', 'impressions', 'clicks', 'ad_orders'):
            c[k] += r[k] or 0
        exp = (r['exposure_id'] or '').strip()
        base = (r['product_name'] or '').split(',')[0].strip()
        pkey = exp or base or '(미상)'
        p = c['products'][pkey]
        p['name'] = p['name'] or base
        p['exposure_id'] = p['exposure_id'] or exp
        for k in ('ad_cost', 'ad_sales', 'clicks'):
            p[k] += r[k] or 0

    ch_qs = CoupangAdChange.objects.all()
    if cupang_id:
        ch_qs = ch_qs.filter(cupang_id=cupang_id)
    ch_by_camp = defaultdict(list)
    for ch in ch_qs:
        ch_by_camp[(ch.campaign_name or '').strip()].append({
            'id': ch.id, 'change_date': str(ch.change_date), 'change_type': ch.change_type,
            'budget_before': ch.budget_before, 'budget_after': ch.budget_after, 'memo': ch.memo,
        })

    # 변경이력에만 있고 아직 광고비 0인 캠페인(신설 등)도 카드로 노출 (빈 집계로 생성)
    for cname in ch_by_camp:
        if cname and cname not in camps:
            camps[cname]  # defaultdict → 0값으로 생성

    out = []
    for cname, c in camps.items():
        prods = sorted(c['products'].values(), key=lambda x: -x['ad_cost'])
        for p in prods:
            p['roas'] = round(p['ad_sales'] / p['ad_cost'] * 100) if p['ad_cost'] else 0
        out.append({
            'campaign_name': cname,
            'ad_cost': c['ad_cost'], 'ad_sales': c['ad_sales'],
            'impressions': c['impressions'], 'clicks': c['clicks'], 'ad_orders': c['ad_orders'],
            'roas': round(c['ad_sales'] / c['ad_cost'] * 100) if c['ad_cost'] else 0,
            'acos': round(c['ad_cost'] / c['ad_sales'] * 100, 1) if c['ad_sales'] else 0,
            'product_count': len(prods),
            'products': prods,
            'change_history': ch_by_camp.get(cname, []),
        })
    out.sort(key=lambda x: -x['ad_cost'])
    return {'range': {'from': str(date_from), 'to': str(date_to)}, 'campaigns': out}


def list_ad_changes(cupang_id=None, limit=200):
    """광고 설정 변경이력 목록 (최신순)."""
    from .models import CoupangAdChange
    qs = CoupangAdChange.objects.all()
    if cupang_id:
        qs = qs.filter(cupang_id=cupang_id)
    return [{
        'id': c.id, 'cupang_id': c.cupang_id, 'change_date': str(c.change_date),
        'change_type': c.change_type, 'campaign_name': c.campaign_name,
        'budget_before': c.budget_before, 'budget_after': c.budget_after,
        'products': c.products or [], 'memo': c.memo,
    } for c in qs[:limit]]


def save_ad_change(data):
    """변경이력 추가/수정. id 있으면 수정."""
    from .models import CoupangAdChange
    fields = dict(
        cupang_id=data.get('cupang_id') or '',
        change_date=data.get('change_date'),
        change_type=data.get('change_type') or 'etc',
        campaign_name=(data.get('campaign_name') or '').strip(),
        budget_before=data.get('budget_before'),
        budget_after=data.get('budget_after'),
        products=data.get('products') or [],
        memo=data.get('memo') or '',
    )
    if data.get('id'):
        CoupangAdChange.objects.filter(pk=data['id']).update(**fields)
        return data['id']
    return CoupangAdChange.objects.create(**fields).id


def delete_ad_change(change_id):
    from .models import CoupangAdChange
    CoupangAdChange.objects.filter(pk=change_id).delete()


def get_coupang_seller_aliases():
    """order DB 06.쿠팡 사업자(seller_alias) 목록."""
    with connections['joacham'].cursor() as cur:
        cur.execute("SELECT DISTINCT seller_alias FROM orders_order WHERE site_name='06.쿠팡' AND seller_alias<>'' ORDER BY seller_alias")
        return [r[0] for r in cur.fetchall()]


# 쿠팡 로그인ID → order DB seller_alias 매핑 (docs '쇼핑몰 계정 관리' 06.쿠팡 별칭 기준)
COUPANG_ALIAS_MAP = {
    'bdshouse': '12바둑이하우스', 'betona': '01비투나', 'bitcom1': '03비트컴',
    'bitic05': '05비트윙', 'bitmind': '02비트마인드', 'compwoow': '11캠핑와우',
    'elike01': '10이처럼', 'erowoo1': '09이로워', 'exansys': '13엑사엔시스',
    'hwss01': '07행원상사', 'joacham': '14조아참', 'joys3763': '04조아스',
    'nainjoy6': '06나인조이', 'nkcms01': '08나경커머스',
}


def get_settlement_verification(date_from='2026-01-01'):
    """정산 검증 — 산술 일관성/API 내부정합/나인조이 로켓/이상치 플래그를 구조화 반환 (UI용)."""
    from django.db.models import Sum
    from .models import CoupangSettlement, CoupangRocketSettlement
    checks = []

    # 1) 산술 일관성: 확정정산 = 마켓 + 로켓
    u = get_unified_settlement(date_from)
    bad = [r['cupang_id'] for r in u['rows'] if r['confirmed'] != r['mp_api_final'] + r['rocket_final']]
    checks.append({'name': '산술 일관성 (확정정산 = 마켓플레이스 + 로켓그로스)',
                   'status': 'pass' if not bad else 'fail',
                   'detail': f"전 계정 {len(u['rows'])}개 일치" if not bad else f"불일치: {', '.join(bad)}"})

    # 2) exansys API 내부정합: 정산액 + 보류 = 정산대상
    exs = CoupangSettlement.objects.filter(account__cupang_id='exansys')
    bad2 = [str(s.settlement_date) for s in exs
            if abs((s.settlement_amount or 0) + (s.last_amount or 0) - (s.settlement_target or 0)) > 1]
    checks.append({'name': 'API 정산 내부정합 (정산액 + 보류30% = 정산대상)',
                   'status': 'pass' if not bad2 else 'warn',
                   'detail': f"exansys {exs.count()}건 모두 정합" if not bad2 else f"불일치 {len(bad2)}건: {', '.join(bad2[:3])}"})

    # 3) 나인조이 로켓그로스 적용 여부
    nj = CoupangRocketSettlement.objects.filter(account__cupang_id='nainjoy6')
    njsum = nj.aggregate(s=Sum('final_amount'))['s'] or 0
    njnz = nj.exclude(final_amount=0).count()
    checks.append({'name': '나인조이 로켓그로스 적용',
                   'status': 'pass' if njsum > 0 else 'warn',
                   'detail': f"{nj.count()}건 중 금액 {njnz}건, 확정 {njsum:,}원 적용됨"})

    # 4) 비트윙 로켓 0원 점검
    b = CoupangRocketSettlement.objects.filter(account__cupang_id='bitic05')
    bgross = b.aggregate(s=Sum('gross_sale'))['s'] or 0
    checks.append({'name': '비트윙(bitic05) 로켓 0원 점검',
                   'status': 'warn',
                   'detail': f"{b.count()}건 전부 매출 0 (gross 합 {bgross}) — 로켓 매출 없음으로 추정, WING 재확인 권장"})

    # 5) API키 미보유 계정
    nokey = [a.cupang_id for a in CoupangApiAccount.objects.all() if not (a.access_key and a.secret_key_enc)]
    checks.append({'name': 'Open API 키 미보유 계정',
                   'status': 'warn' if nokey else 'pass',
                   'detail': f"{len(nokey)}개 — 마켓플레이스 실정산 미확인(대장 추정만): {', '.join(nokey)}" if nokey else "전 계정 키 보유"})

    summary = {'pass': sum(1 for c in checks if c['status'] == 'pass'),
               'warn': sum(1 for c in checks if c['status'] == 'warn'),
               'fail': sum(1 for c in checks if c['status'] == 'fail')}
    return {'checks': checks, 'summary': summary, 'date_from': date_from}


def get_unified_settlement(date_from=None, date_to=None):
    """전 계정 통합 정산 — 로켓그로스(WING 실정산) + 마켓플레이스(API 실정산 or 대장 추정).
    '확정 정산' = 실제 받은 돈(API + 로켓 WING). '대장'은 참고(내 장부 총매출)."""
    from django.db.models import Sum
    from .models import CoupangRocketSettlement
    RET = ("order_status NOT LIKE '%%취소%%' AND order_status NOT LIKE '%%반품%%' "
           "AND order_status NOT LIKE '%%환불%%' AND order_status NOT LIKE '%%교환%%'")
    rows = []
    tot = {'mp_api': 0, 'rocket': 0, 'confirmed': 0, 'ledger': 0}
    for acc in CoupangApiAccount.objects.all().order_by('cupang_id'):
        cid = acc.cupang_id
        has_key = bool(acc.access_key and acc.secret_key_enc)
        # ① 마켓플레이스 — API 실정산
        mq = CoupangSettlement.objects.filter(account=acc)
        if date_from:
            mq = mq.filter(settlement_date__gte=date_from)
        if date_to:
            mq = mq.filter(settlement_date__lte=date_to)
        mp_api = mq.aggregate(s=Sum('final_amount'))['s'] or 0
        mp_cnt = mq.count()
        # ② 로켓그로스 — WING 크롤 실정산
        rq = CoupangRocketSettlement.objects.filter(account=acc)
        if date_from:
            rq = rq.filter(settlement_date__gte=date_from)
        if date_to:
            rq = rq.filter(settlement_date__lte=date_to)
        rocket = rq.aggregate(s=Sum('final_amount'))['s'] or 0
        rg_cnt = rq.count()
        # ③ 대장(참고) — order DB
        alias = COUPANG_ALIAS_MAP.get(cid)
        ledger = 0
        ledger_cnt = 0
        if alias:
            where = "site_name='06.쿠팡' AND seller_alias=%s AND " + RET
            params = [alias]
            if date_from:
                where += " AND order_date>=%s"; params.append(date_from)
            if date_to:
                where += " AND order_date<=%s"; params.append(str(date_to) + ' 23:59:59')
            with connections['joacham'].cursor() as c:
                c.execute(f"SELECT COUNT(*), COALESCE(SUM(settlement_price),0) FROM orders_order WHERE {where}", params)
                ledger_cnt, ledger = c.fetchone()
                ledger = int(ledger or 0)
        confirmed = mp_api + rocket   # 실제 확정 정산(받은 돈)
        if mp_cnt == 0 and rg_cnt == 0 and ledger_cnt == 0:
            continue  # 데이터 전혀 없는 계정 제외
        rows.append({
            'cupang_id': cid, 'account_name': acc.account_name or '', 'alias': alias or '',
            'has_api_key': has_key,
            'mp_api_final': mp_api, 'mp_cnt': mp_cnt,
            'mp_source': 'API실정산' if has_key else ('대장추정' if ledger else '-'),
            'rocket_final': rocket, 'rg_cnt': rg_cnt,
            'ledger': ledger, 'ledger_cnt': ledger_cnt,
            'confirmed': confirmed,
            # 검증: API키 있으면 확정, 없으면 대장만(추정)
            'verified': has_key or rg_cnt > 0,
        })
        tot['mp_api'] += mp_api; tot['rocket'] += rocket
        tot['confirmed'] += confirmed; tot['ledger'] += ledger
    rows.sort(key=lambda r: -(r['confirmed'] or r['ledger']))
    return {'rows': rows, 'totals': tot}


def get_settlements(account_id=None, year_month=None):
    qs = CoupangSettlement.objects.select_related('account').all()
    if account_id:
        qs = qs.filter(account_id=account_id)
    if year_month:
        qs = qs.filter(revenue_ym=year_month)
    rows = [{
        'cupang_id': s.account.cupang_id,
        'settlement_type': s.settlement_type,
        'settlement_date': s.settlement_date.strftime('%Y-%m-%d') if s.settlement_date else None,
        'revenue_ym': s.revenue_ym,
        'recognition_from': s.recognition_from.strftime('%Y-%m-%d') if s.recognition_from else None,
        'recognition_to': s.recognition_to.strftime('%Y-%m-%d') if s.recognition_to else None,
        'total_sale': s.total_sale, 'service_fee': s.service_fee,
        'settlement_target': s.settlement_target, 'settlement_amount': s.settlement_amount,
        'deduction_amount': s.deduction_amount, 'last_amount': s.last_amount,
        'final_amount': s.final_amount, 'status': s.status,
        'bank_name': s.bank_name, 'bank_account': s.bank_account, 'bank_holder': s.bank_holder,
    } for s in qs]
    totals = {
        'count': len(rows),
        'total_sale': sum(r['total_sale'] for r in rows),
        'service_fee': sum(r['service_fee'] for r in rows),
        'final_amount': sum(r['final_amount'] for r in rows),
    }
    return {'rows': rows, 'totals': totals}


# ── 단일 옵션 재고 조회 ──
def get_stock_amount(account: CoupangApiAccount, vendor_item_id: str, max_retries: int = 5):
    """쿠팡 Open API 재고조회 → (sale_price, amount_in_stock, error)
    error 가 None 이면 성공."""
    vid = clean_token(vendor_item_id)
    if not vid.isnumeric():
        return None, None, f"잘못된 옵션 ID: {vendor_item_id!r}"

    secret_key = decrypt_secret(account.secret_key_enc)
    access_key = account.access_key
    vendor_id = account.vendor_id

    retry_delay = 2
    for attempt in range(max_retries):
        datetime_str = time.strftime('%y%m%dT%H%M%SZ', time.gmtime())  # 쿠팡 CEA 스펙: 2자리 연도 UTC
        method = "GET"
        path = INVENTORY_PATH.format(vid=vid)
        message = f"{datetime_str}{method}{path}"
        signature = hmac.new(secret_key.encode('utf-8'), message.encode('utf-8'), hashlib.sha256).hexdigest()
        authorization = (
            f"CEA algorithm=HmacSHA256, access-key={access_key}, "
            f"signed-date={datetime_str}, signature={signature}"
        )

        req = urllib.request.Request(API_HOST + path)
        req.add_header("Content-Type", "application/json;charset=UTF-8")
        req.add_header("Authorization", authorization)
        if vendor_id:
            req.add_header("X-Requested-By", vendor_id)

        ctx = ssl.create_default_context()
        ctx.check_hostname = False
        ctx.verify_mode = ssl.CERT_NONE

        try:
            with urllib.request.urlopen(req, context=ctx, timeout=30) as resp:
                charset = resp.headers.get_content_charset() or 'utf-8'
                data = json.loads(resp.read().decode(charset))
                if data.get("code") == "SUCCESS":
                    d = data.get("data", {})
                    return d.get("salePrice"), d.get("amountInStock"), None
                return None, None, f"API 응답 실패: {data.get('message') or data}"
        except urllib.error.HTTPError as e:
            body = ''
            try:
                body = e.read().decode()
            except Exception:
                pass
            if e.code == 429:  # rate limit → 지수 백오프 재시도
                time.sleep(retry_delay)
                retry_delay *= 2
                continue
            hint = ''
            if e.code == 401:
                hint = " (401: 키/계정 매칭, vendorId, UTC 시각 확인)"
            return None, None, f"HTTP {e.code} {e.reason}{hint}: {body[:200]}"
        except urllib.error.URLError as e:
            return None, None, f"URL 오류: {e.reason}"
        except Exception as e:
            return None, None, f"예외: {e}"
    return None, None, "최대 재시도 실패(429)"


# ── 재고 로그 저장 + 상품 최신값 갱신 ──
def record_inventory(product: CoupangRocketProduct, sale_price, stock):
    prev = product.last_stock  # 직전 측정값 (갱신 전)
    delta = None
    if prev is not None and stock is not None:
        delta = stock - prev   # +입고 / -판매
    CoupangInventoryLog.objects.create(
        vendor_item_id=product.vendor_item_id,
        sale_price=sale_price,
        stock=stock,
        prev_stock=prev,
        delta=delta,
    )
    # 가격 변동 감지 → 이력 저장
    if (sale_price is not None and product.last_price is not None
            and sale_price != product.last_price):
        CoupangPriceChange.objects.create(
            vendor_item_id=product.vendor_item_id,
            old_price=product.last_price,
            new_price=sale_price,
        )
    product.last_price = sale_price
    product.last_stock = stock
    product.last_checked_at = timezone.now()
    product.save(update_fields=['last_price', 'last_stock', 'last_checked_at'])
    # 입고 예정 자동 매칭 (증가 이벤트가 예정수량에 근접하면 입고로 분류)
    if delta and delta > 0:
        try:
            er = _match_expected_restock(product.vendor_item_id, delta)
            if er is None:
                _auto_first_restock(product, prev, stock)   # 입고예정 없이 0→증가 = 첫입고
        except Exception:
            pass
    _evaluate_alarm(product, stock)


def _auto_first_restock(product, prev_stock, stock):
    """입고예정 없이 재고가 0(또는 최초)에서 증가하면 '첫입고'로 자동 기록.
    윈도우 내 모든 증가분을 누적해 총입고에 반영 + 입고완료 팝업 노출."""
    from .models import CoupangExpectedRestock, CoupangInventoryLog
    from datetime import timedelta
    if not (stock and stock > 0):
        return None
    if prev_stock not in (None, 0):
        return None
    vid = product.vendor_item_id
    if CoupangExpectedRestock.objects.filter(
            vendor_item_id=vid, status__in=['pending', 'matched']).exists():
        return None
    first_inc = (CoupangInventoryLog.objects
                 .filter(vendor_item_id=vid, delta__gt=0)
                 .order_by('checked_at').values_list('checked_at', flat=True).first())
    reg = (first_inc - timedelta(seconds=1)) if first_inc else (timezone.now() - timedelta(minutes=1))
    er = CoupangExpectedRestock.objects.create(
        vendor_item_id=vid, expected_quantity=0, window_days=30,
        status='pending', memo='첫입고 자동감지',
    )
    # registered_at 은 auto_now_add 라 create 로 못 박음 → update 로 강제(첫 증가 직전 시점)
    CoupangExpectedRestock.objects.filter(id=er.id).update(registered_at=reg)
    er.refresh_from_db()
    inc_logs = _er_window_increase_logs(er)
    if inc_logs:
        _apply_er_restock(er, inc_logs)
        return er
    er.delete()
    return None


def _er_window_increase_logs(er):
    """입고예정(er) 등록 이후 윈도우 내 양의 증가 로그 전체.
    첫입고는 모든 증분이 입고이므로 +1/+2 도 포함(누적이 실재고와 일치)."""
    from .models import CoupangInventoryLog
    from datetime import timedelta
    if not er.registered_at:
        return []
    window_end = er.registered_at + timedelta(days=er.window_days)
    return list(CoupangInventoryLog.objects
                .filter(vendor_item_id=er.vendor_item_id, delta__gt=0,
                        checked_at__gte=er.registered_at, checked_at__lte=window_end)
                .order_by('checked_at'))


def _apply_er_restock(er, inc_logs, status='matched'):
    """입고예정에 누적 입고량 반영 + 기여 로그 입고표시 + 일별 재집계."""
    from .models import CoupangInventoryLog
    cum = sum(l.delta for l in inc_logs)
    CoupangInventoryLog.objects.filter(id__in=[l.id for l in inc_logs]).update(marked_restock=True)
    last = inc_logs[-1]
    er.status = status
    er.matched_at = last.checked_at
    er.matched_qty = cum            # 전체 누적 = 실입고량
    er.matched_log_id = last.id
    er.save(update_fields=['status', 'matched_at', 'matched_qty', 'matched_log_id'])
    for d in {l.checked_at.date() for l in inc_logs}:
        rebuild_daily_sales(er.vendor_item_id, only_date=d)
    return cum


def _match_expected_restock(vid, delta):
    """재고 증가를 입고예정과 매칭.
    ① pending: 등록 이후 윈도우 내 누적 증가분이 예정수량의 50% 이상이면 입고 인정.
    ② 이미 matched: 이후 추가 입고분까지 matched_qty 계속 갱신(쿠팡 입고는 나눠 들어옴).
    matched_qty = 윈도우 내 전체 누적 증가 = 실입고량(총입고에 그대로 반영)."""
    from .models import CoupangExpectedRestock
    from datetime import timedelta
    now = timezone.now()
    matched_er = None
    for er in CoupangExpectedRestock.objects.filter(
            vendor_item_id=vid, status__in=['pending', 'matched']).order_by('registered_at'):
        if er.status == 'pending' and er.registered_at and (now - er.registered_at).days > er.window_days:
            er.status = 'expired'
            er.save(update_fields=['status'])
            continue
        inc_logs = _er_window_increase_logs(er)
        if not inc_logs:
            continue
        cum = sum(l.delta for l in inc_logs)
        if er.status == 'matched':
            # 추가 입고분 누적 갱신
            if cum != (er.matched_qty or 0):
                _apply_er_restock(er, inc_logs)
            matched_er = matched_er or er
        elif cum >= RESTOCK_MIN_DELTA:
            # 실입고 감지(주문취소 +1/+2 블립보다 큼) → 즉시 입고처리, 이후 누적 갱신.
            # 예정수량 미달이어도 들어온 만큼 총입고에 반영(부분입고 표시).
            _apply_er_restock(er, inc_logs)
            matched_er = matched_er or er
    return matched_er


def get_restock_detail(vid):
    """옵션(vid)의 입고 내역 — 수동입고 + 자동매칭 입고예정. 날짜·수량·구분.
    총입고 숫자 클릭 모달용."""
    from .models import CoupangExpectedRestock, CoupangRestock
    out = []
    for r in CoupangRestock.objects.filter(vendor_item_id=vid).order_by('restock_date'):
        out.append({
            'date': r.restock_date.strftime('%Y-%m-%d') if r.restock_date else '',
            'time': '',
            'qty': r.quantity,
            'source': '수동입고' if r.source == 'manual' else (r.source or '입고'),
            'memo': r.memo or '',
        })
    for er in CoupangExpectedRestock.objects.filter(vendor_item_id=vid, status='matched').order_by('matched_at'):
        out.append({
            'date': er.matched_at.strftime('%Y-%m-%d') if er.matched_at else '',
            'time': er.matched_at.strftime('%H:%M') if er.matched_at else '',
            'qty': er.matched_qty or 0,
            'source': '자동감지',
            'memo': f"예정 {er.expected_quantity}개" + (f" · {er.memo}" if er.memo else ''),
        })
    out.sort(key=lambda x: (x['date'], x['time']))
    total = sum(x['qty'] for x in out)
    return {'rows': out, 'total': total}


def _evaluate_alarm(product: CoupangRocketProduct, stock):
    """알람 활성 상품: 재고가 임계 이하로 내려가면 텔레그램 1회 발송,
    임계 초과로 회복되면 알람상태 해제(다음 하락 시 재발송)."""
    if not product.alarm_enabled or stock is None:
        return
    threshold = product.alarm_threshold or 10
    if stock <= threshold and not product.alarm_notified:
        _send_stock_alarm(product, stock, threshold)
        product.alarm_notified = True
        product.save(update_fields=['alarm_notified'])
    elif stock > threshold and product.alarm_notified:
        product.alarm_notified = False
        product.save(update_fields=['alarm_notified'])


def _send_stock_alarm(product: CoupangRocketProduct, stock, threshold):
    from . import telegram
    name = f"{product.product_name or ''} {product.option_name or ''}".strip() or product.vendor_item_id
    if stock <= 0:
        head = "🚨 [쿠팡로켓 품절]"
        body = f"{name}\n옵션ID {product.vendor_item_id}\n재고 0개 — 즉시 입고 필요!"
    else:
        head = "⚠️ [쿠팡로켓 재고부족]"
        body = f"{name}\n옵션ID {product.vendor_item_id}\n재고 {stock}개 (임계 {threshold}개 이하)"
    try:
        telegram.send_telegram(f"{head}\n{body}\n계정: {product.account.cupang_id}")
    except Exception:
        pass


def backfill_deltas(vendor_item_id: str = None):
    """기존 로그의 prev_stock/delta 를 연속 측정 순서대로 재계산."""
    vids = ([vendor_item_id] if vendor_item_id
            else CoupangInventoryLog.objects.values_list('vendor_item_id', flat=True).distinct())
    n = 0
    for vid in vids:
        prev = None
        logs = CoupangInventoryLog.objects.filter(vendor_item_id=vid).order_by('checked_at')
        for log in logs:
            new_delta = (log.stock - prev) if (prev is not None and log.stock is not None) else None
            if log.prev_stock != prev or log.delta != new_delta:
                log.prev_stock = prev
                log.delta = new_delta
                log.save(update_fields=['prev_stock', 'delta'])
                n += 1
            prev = log.stock
    return n


# ── 일별 판매량/입고량 계산 (연속 측정값 구간 차분 → 입고 자동보정) ──
def _aggregate_day(vid, d):
    """특정 옵션·날짜의 (sold, restock, first_stock, last_stock) 산출.

    저장된 delta(+증가/-감소) + marked_restock 분류 사용:
    - 감소(delta<0) → 총판매(gross_sold)
    - 증가(delta>0): 기본 '주문취소' → gross_sold 에서 차감(순판매↓).
                     marked_restock=True 인 건만 '입고'로 분리(순판매 영향 없음).
    - 순판매(sold) = max(0, gross_sold - 취소합)
    """
    day_logs = list(
        CoupangInventoryLog.objects
        .filter(vendor_item_id=vid, checked_at__date=d)
        .exclude(stock__isnull=True)
        .order_by('checked_at')
        .values_list('stock', 'delta', 'marked_restock')
    )
    if not day_logs:
        return None
    first_stock = day_logs[0][0]
    last_stock = day_logs[-1][0]
    gross_sold = cancel = restock = 0
    for stock, delta, is_restock in day_logs:
        if delta is None:
            continue
        if delta < 0:
            if -delta <= ABNORMAL_DROP:        # 비정상 대량감소(품절/조정/입력오류) 제외
                gross_sold += -delta
        elif delta > 0:
            if is_restock:
                restock += delta      # 입고 (수동 표시)
            else:
                cancel += delta       # 주문취소 (기본)
    sold = max(0, gross_sold - cancel)
    return sold, restock, first_stock, last_stock


def rebuild_daily_sales(vendor_item_id: str = None, only_date=None):
    """inventory_log 로부터 일별 판매량/입고량을 재집계.
    only_date 지정 시 그 날짜만(잦은 실행용), 미지정 시 전체 날짜."""
    qs = CoupangRocketProduct.objects.all()
    if vendor_item_id:
        qs = qs.filter(vendor_item_id=vendor_item_id)

    count = 0
    for product in qs:
        vid = product.vendor_item_id
        if only_date is not None:
            dates = [only_date]
        else:
            dates = CoupangInventoryLog.objects.filter(vendor_item_id=vid).dates('checked_at', 'day')
        for d in dates:
            agg = _aggregate_day(vid, d)
            if agg is None:
                continue
            sold, restock, first_stock, last_stock = agg
            CoupangDailySales.objects.update_or_create(
                vendor_item_id=vid, date=d,
                defaults={
                    'sold_quantity': sold,
                    'restock_quantity': restock,
                    'first_stock': first_stock,
                    'last_stock': last_stock,
                },
            )
            count += 1
    return count


# ── 재고 증가 이벤트 (주문취소/입고 분류) ──
def get_increase_events(vendor_item_id=None, days=30):
    """재고 증가(delta>0) 이벤트 목록. 기본 분류는 주문취소, marked_restock=True 면 입고."""
    from datetime import timedelta, date as _date
    since = _date.today() - timedelta(days=days)
    qs = (CoupangInventoryLog.objects
          .filter(delta__gt=0, checked_at__date__gte=since)
          .order_by('-checked_at'))
    if vendor_item_id:
        qs = qs.filter(vendor_item_id=vendor_item_id)
    prods = {p.vendor_item_id: p for p in CoupangRocketProduct.objects.all()}
    rows = []
    for lg in qs[:500]:
        p = prods.get(lg.vendor_item_id)
        rows.append({
            'id': lg.id,
            'vendor_item_id': lg.vendor_item_id,
            'product_name': p.product_name if p else '',
            'option_name': p.option_name if p else '',
            'checked_at': lg.checked_at.strftime('%Y-%m-%d %H:%M'),
            'date': lg.checked_at.strftime('%Y-%m-%d'),
            'prev_stock': lg.prev_stock,
            'stock': lg.stock,
            'delta': lg.delta,
            'marked_restock': lg.marked_restock,
        })
    return rows


def set_increase_kind(log_id, is_restock):
    """증가 이벤트를 입고(True)/주문취소(False)로 분류 + 해당 날짜 일별판매 재집계."""
    lg = CoupangInventoryLog.objects.filter(pk=log_id, delta__gt=0).first()
    if not lg:
        return None
    lg.marked_restock = bool(is_restock)
    lg.save(update_fields=['marked_restock'])
    rebuild_daily_sales(lg.vendor_item_id, only_date=lg.checked_at.date())
    return {'id': lg.id, 'marked_restock': lg.marked_restock}


# ── 입고 예정 (자동 매칭) ──
def register_expected_restock(vid, quantity, window_days=7, memo=''):
    from .models import CoupangExpectedRestock
    er = CoupangExpectedRestock.objects.create(
        vendor_item_id=vid, expected_quantity=int(quantity),
        window_days=int(window_days or 7), memo=memo or '')
    return er


def delete_expected_restock(er_id):
    from .models import CoupangExpectedRestock
    CoupangExpectedRestock.objects.filter(pk=er_id).delete()


def list_expected_restocks(vid=None):
    """입고예정 목록 (만료 자동반영). 상품정보 포함."""
    from .models import CoupangExpectedRestock
    now = timezone.now()
    qs = CoupangExpectedRestock.objects.all()
    if vid:
        qs = qs.filter(vendor_item_id=vid)
    prods = {p.vendor_item_id: p for p in CoupangRocketProduct.objects.all()}
    out = []
    for er in qs:
        st = er.status
        if st == 'pending' and er.registered_at and (now - er.registered_at).days > er.window_days:
            st = 'expired'
            CoupangExpectedRestock.objects.filter(pk=er.id).update(status='expired')
        p = prods.get(er.vendor_item_id)
        out.append({
            'id': er.id,
            'vendor_item_id': er.vendor_item_id,
            'product_id': p.id if p else None,
            'has_image': bool(p.image_file) if p else False,
            'product_name': p.product_name if p else '',
            'option_name': p.option_name if p else '',
            'expected_quantity': er.expected_quantity,
            'window_days': er.window_days,
            'status': st,
            'matched_at': er.matched_at.strftime('%Y-%m-%d %H:%M') if er.matched_at else None,
            'matched_qty': er.matched_qty,
            'memo': er.memo,
            'registered_at': er.registered_at.strftime('%Y-%m-%d %H:%M') if er.registered_at else None,
        })
    return out


def get_restock_summary(vid=None):
    """옵션별 입고예정수량(pending) + 총입고수량.
    ★ 총입고 = 현재고 + 누적판매 (재고 보존법칙). 선입고 미등록으로 들어온 입고도 자동 반영
       (과거엔 CoupangRestock+matched만 합산 → 선입고 안 걸고 입고하면 누락되는 버그.
        예: 퍼플 현재고64인데 총입고10, 팔뚝 블랙1 현재고543인데 총입고396)."""
    from django.db.models import Sum
    from .models import CoupangExpectedRestock, CoupangRocketProduct, CoupangDailySales
    # 현재고 (API 최신값)
    prods = CoupangRocketProduct.objects.all()
    if vid:
        prods = prods.filter(vendor_item_id=vid)
    stock_map = {v: (s or 0) for v, s in prods.values_list('vendor_item_id', 'last_stock')}
    # 누적판매 (일별판매 합)
    ds = CoupangDailySales.objects.all()
    if vid:
        ds = ds.filter(vendor_item_id=vid)
    sold_map = {}
    for r in ds.values('vendor_item_id').annotate(s=Sum('sold_quantity')):
        sold_map[r['vendor_item_id']] = r['s'] or 0
    # 입고예정수량 = pending 합 (선입고)
    exp = CoupangExpectedRestock.objects.filter(status='pending')
    if vid:
        exp = exp.filter(vendor_item_id=vid)
    pend_map = {}
    for r in exp.values('vendor_item_id').annotate(t=Sum('expected_quantity')):
        pend_map[r['vendor_item_id']] = r['t'] or 0
    keys = set(stock_map) | set(sold_map) | set(pend_map)
    return {k: {'pending_qty': pend_map.get(k, 0),
                'total_restock': stock_map.get(k, 0) + sold_map.get(k, 0)} for k in keys}


# ── 입고(재입고) ──
def add_restock(product: CoupangRocketProduct, restock_date, quantity: int, memo: str = '', source: str = 'manual'):
    """입고 이력 기록 + 해당 날짜에 재고 증가(+quantity)를 시계열에 반영.
    restock_date 시점에 backdated inventory_log(delta=+quantity) 를 삽입해
    일별 입고/판매 집계와 그래프에 자동 반영. 추후 쿠팡 입고API 가 source='coupang' 으로 호출."""
    from datetime import datetime, time as dtime, date as _date

    if isinstance(restock_date, str):
        restock_date = datetime.strptime(restock_date, '%Y-%m-%d').date()
    quantity = int(quantity)

    rs = CoupangRestock.objects.create(
        vendor_item_id=product.vendor_item_id,
        restock_date=restock_date,
        quantity=quantity,
        source=source,
        memo=memo or '',
    )

    # 입고 시점 직전 재고 → +quantity 로 backdated 측정 삽입
    when = datetime.combine(restock_date, dtime(12, 0))
    prev = (
        CoupangInventoryLog.objects
        .filter(vendor_item_id=product.vendor_item_id, checked_at__lt=when)
        .exclude(stock__isnull=True)
        .order_by('-checked_at')
        .values_list('stock', flat=True)
        .first()
    ) or 0
    new_stock = prev + quantity
    log = CoupangInventoryLog.objects.create(
        vendor_item_id=product.vendor_item_id,
        sale_price=product.last_price,
        stock=new_stock,
        prev_stock=prev,
        delta=quantity,
    )
    CoupangInventoryLog.objects.filter(id=log.id).update(checked_at=when)

    # NOTE: 과거엔 오늘 입고 시 product.last_stock 을 prev+quantity 로 즉시 올렸으나,
    #   다음 폴링이 (아직 안 들어온) 실제재고를 읽어 delta=실재고-부풀린last_stock = 음수(유령 판매)를
    #   만들었음(시간대별 판매 오집계). last_stock 은 실제 재고만 반영하도록 갱신하지 않는다.
    #   (입고 자체는 위 backdated 로그 + rebuild_daily_sales 로 그래프/집계에 반영됨)

    rebuild_daily_sales(vendor_item_id=product.vendor_item_id)
    return rs


def remove_restock(restock: CoupangRestock):
    """입고 이력 삭제 + 그때 삽입한 backdated 재고측정 제거 + 재집계."""
    from datetime import datetime, time as dtime
    vid = restock.vendor_item_id
    when = datetime.combine(restock.restock_date, dtime(12, 0))
    # add_restock 이 만든 backdated 측정(같은 날 12:00, delta=+quantity) 제거
    log = (
        CoupangInventoryLog.objects
        .filter(vendor_item_id=vid, checked_at=when, delta=restock.quantity)
        .order_by('-id')
        .first()
    )
    if log:
        log.delete()
    restock.delete()
    backfill_deltas(vid)
    rebuild_daily_sales(vendor_item_id=vid)
    # 현재 재고를 최신 실측값으로 복원
    latest = (
        CoupangInventoryLog.objects.filter(vendor_item_id=vid)
        .exclude(stock__isnull=True).order_by('-checked_at')
        .values_list('stock', flat=True).first()
    )
    CoupangRocketProduct.objects.filter(vendor_item_id=vid).update(last_stock=latest)


# ── 상품 상세 히스토리 (30일 판매 + 가격 변동) ──
def get_product_history(vendor_item_id: str, days: int = 30):
    from datetime import date as _date, datetime, timedelta
    from collections import OrderedDict

    today = _date.today()
    start = today - timedelta(days=days - 1)
    start_dt = datetime.combine(start, datetime.min.time())

    # 일별 판매/입고
    sales = {
        r.date: r for r in
        CoupangDailySales.objects.filter(vendor_item_id=vendor_item_id, date__gte=start)
    }
    # 일별 가격(그날 마지막 측정가) — inventory_log 에서
    price_by_day = OrderedDict()
    for stock, price, checked_at in (
        CoupangInventoryLog.objects
        .filter(vendor_item_id=vendor_item_id, checked_at__gte=start_dt)
        .exclude(sale_price__isnull=True)
        .order_by('checked_at')
        .values_list('stock', 'sale_price', 'checked_at')
    ):
        price_by_day[checked_at.date()] = price

    # 시작 이전 마지막 가격(캐리)
    last_price = (
        CoupangInventoryLog.objects
        .filter(vendor_item_id=vendor_item_id, checked_at__lt=start_dt)
        .exclude(sale_price__isnull=True)
        .order_by('-checked_at')
        .values_list('sale_price', flat=True)
        .first()
    )

    series = []
    carry = last_price
    for i in range(days):
        d = start + timedelta(days=i)
        if d in price_by_day:
            carry = price_by_day[d]
        s = sales.get(d)
        series.append({
            'date': d.strftime('%m-%d'),
            'full_date': d.strftime('%Y-%m-%d'),
            'sold': s.sold_quantity if s else 0,
            'restock': s.restock_quantity if s else 0,
            'stock': s.last_stock if s else None,
            'price': carry,
        })

    price_changes = [
        {
            'changed_at': c.changed_at.strftime('%Y-%m-%d %H:%M'),
            'old_price': c.old_price,
            'new_price': c.new_price,
        }
        for c in CoupangPriceChange.objects.filter(vendor_item_id=vendor_item_id).order_by('-changed_at')[:50]
    ]

    restocks = [
        {
            'id': r.id,
            'restock_date': r.restock_date.strftime('%Y-%m-%d'),
            'quantity': r.quantity,
            'source': r.source,
            'memo': r.memo,
        }
        for r in CoupangRestock.objects.filter(vendor_item_id=vendor_item_id).order_by('-restock_date', '-id')[:50]
    ]

    return {
        'vendor_item_id': vendor_item_id,
        'days': days,
        'series': series,
        'price_changes': price_changes,
        'restocks': restocks,
        'total_sold': sum(p['sold'] for p in series),
        'total_restock': sum(r['quantity'] for r in restocks),
    }


# ── 대시보드 통계 ──
def get_dashboard_stats(account_id: int = None, pattern_days: int = 30, date_str: str = None):
    """상품별/옵션별 오늘·7일·30일 판매량 + 매출 + 선택일 시간대별 매출 그래프.
    pattern_days: 시간대별 판매 패턴 집계 기간(7 또는 30).
    date_str: 매출그래프/시간대별 상품리스트가 표시할 날짜(YYYY-MM-DD). 미지정 시 오늘."""
    from datetime import date as _date, datetime, timedelta
    from collections import defaultdict

    today = _date.today()
    # 날짜 네비게이션 대상일(매출그래프 + 시간대별 상품리스트만 영향)
    sel_date = today
    if date_str:
        try:
            d = datetime.strptime(date_str, '%Y-%m-%d').date()
            if d <= today:
                sel_date = d
        except (TypeError, ValueError):
            pass
    yesterday = today - timedelta(days=1)
    week_start = today - timedelta(days=6)
    month_start = today - timedelta(days=29)
    pattern_days = 7 if pattern_days == 7 else 30
    pattern_start = today - timedelta(days=pattern_days - 1)

    qs = CoupangRocketProduct.objects.filter(is_active=True).select_related('account')
    if account_id:
        qs = qs.filter(account_id=account_id)
    products = list(qs)
    vids = [p.vendor_item_id for p in products]

    # 일별 판매 집계 (최근 30일)
    daily = CoupangDailySales.objects.filter(vendor_item_id__in=vids, date__gte=month_start)
    sold_today = defaultdict(int)
    sold_yesterday = defaultdict(int)
    sold_week = defaultdict(int)
    sold_month = defaultdict(int)
    sold_wkend = defaultdict(int)   # 30일 주말(토·일) 판매
    sold_wkday = defaultdict(int)   # 30일 평일 판매
    for r in daily.values('vendor_item_id', 'date', 'sold_quantity'):
        vid = r['vendor_item_id']
        q = r['sold_quantity'] or 0
        sold_month[vid] += q
        if r['date'].weekday() >= 5:
            sold_wkend[vid] += q
        else:
            sold_wkday[vid] += q
        if r['date'] >= week_start:
            sold_week[vid] += q
        if r['date'] == today:
            sold_today[vid] += q
        if r['date'] == yesterday:
            sold_yesterday[vid] += q

    restock_sum = get_restock_summary()  # 옵션별 입고예정/총입고
    options = []
    for p in products:
        vid = p.vendor_item_id
        price = p.last_price or 0
        t_qty = sold_today[vid]
        rs = restock_sum.get(vid, {})
        options.append({
            'id': p.id,
            'vendor_item_id': vid,
            'seller_product_id': p.seller_product_id,
            'has_image': bool(p.image_file),
            'product_name': p.product_name,
            'option_name': p.option_name,
            'last_price': p.last_price,
            'last_stock': p.last_stock,
            'today_qty': t_qty,
            'yesterday_qty': sold_yesterday[vid],
            'today_amount': t_qty * price,
            'week_qty': sold_week[vid],
            'month_qty': sold_month[vid],
            'pending_restock': rs.get('pending_qty', 0),    # 입고 예정수량(선입고)
            'total_restock': rs.get('total_restock', 0),    # 총입고수량(=현재고+누적판매)
            # 입고필요: 현재고 < 1달판매량 AND 선입고 미등록 → 곧 품절(입고 필요)
            'restock_needed': bool(sold_month[vid] > (p.last_stock or 0) and rs.get('pending_qty', 0) == 0),
        })

    # 30일 윈도우의 주말/평일 일수 (일평균 계산용)
    _wkend_days = sum(1 for i in range(30) if (month_start + timedelta(days=i)).weekday() >= 5)
    _wkday_days = 30 - _wkend_days

    # ── 상품별(노출상품ID 묶음) 집계 → 베스트 상품 ──
    groups = {}
    for o in options:
        key = o['seller_product_id'] or f"_{o['id']}"
        g = groups.get(key)
        if g is None:
            g = groups[key] = {
                'product_key': key,
                'product_name': o['product_name'],
                'image_id': o['id'] if o['has_image'] else None,
                'option_count': 0,
                'today_qty': 0, 'week_qty': 0, 'month_qty': 0, 'today_amount': 0,
                'wkend_qty': 0, 'wkday_qty': 0,
            }
        if not g['product_name'] and o['product_name']:
            g['product_name'] = o['product_name']
        if g['image_id'] is None and o['has_image']:
            g['image_id'] = o['id']
        g['option_count'] += 1
        g['today_qty'] += o['today_qty']
        g['week_qty'] += o['week_qty']
        g['month_qty'] += o['month_qty']
        g['today_amount'] += o['today_amount']
        g['wkend_qty'] += sold_wkend[o['vendor_item_id']]
        g['wkday_qty'] += sold_wkday[o['vendor_item_id']]

    def _verdict(g):
        """주말형/평일형 판정 (일평균 ±20% 기준)."""
        we = round(g['wkend_qty'] / _wkend_days, 1) if _wkend_days else 0
        wd = round(g['wkday_qty'] / _wkday_days, 1) if _wkday_days else 0
        if wd and we >= wd * 1.2:
            v, ratio = '주말형', round(we / wd, 2) if wd else 0
        elif we and wd >= we * 1.2:
            v, ratio = '평일형', round(wd / we, 2) if we else 0
        else:
            v, ratio = '고른편', 1.0
        return {'verdict': v, 'ratio': ratio, 'weekend_avg': we, 'weekday_avg': wd}

    def _rank_products(key, n=5):
        ranked = sorted([g for g in groups.values() if g[key] > 0], key=lambda g: -g[key])
        return [{**g, **_verdict(g)} for g in ranked[:n]]

    best_daily = _rank_products('today_qty')
    best_weekly = _rank_products('week_qty')
    best_monthly = _rank_products('month_qty')

    # ── 상품별 색상 + 일별 판매추이(30일) + 오늘 상품별(파이) ──
    PALETTE = ['#e44232', '#0074e9', '#16a34a', '#f59e0b', '#8b5cf6', '#ec4899', '#06b6d4', '#64748b']
    group_list = list(groups.values())
    products_meta = []
    for i, g in enumerate(group_list):
        g['_name'] = g['product_name'] or g['product_key']
        g['_color'] = PALETTE[i % len(PALETTE)]
        products_meta.append({'name': g['_name'], 'color': g['_color']})

    vid_to_pname = {}
    for o in options:
        key = o['seller_product_id'] or f"_{o['id']}"
        vid_to_pname[o['vendor_item_id']] = groups[key]['_name']

    day_prod_qty = defaultdict(lambda: defaultdict(int))
    day_prod_amt = defaultdict(lambda: defaultdict(int))
    price_map = {o['vendor_item_id']: (o['last_price'] or 0) for o in options}
    for r in daily.values('vendor_item_id', 'date', 'sold_quantity'):
        pn = vid_to_pname.get(r['vendor_item_id'])
        if pn is None:
            continue
        q = r['sold_quantity'] or 0
        day_prod_qty[r['date']][pn] += q
        day_prod_amt[r['date']][pn] += q * price_map.get(r['vendor_item_id'], 0)

    product_daily = []
    for i in range(30):
        d = month_start + timedelta(days=i)
        row = {'date': d.strftime('%m-%d'), 'full_date': d.strftime('%Y-%m-%d')}
        for m in products_meta:
            row[m['name']] = day_prod_qty.get(d, {}).get(m['name'], 0)
            row[m['name'] + '__amt'] = day_prod_amt.get(d, {}).get(m['name'], 0)
        product_daily.append(row)

    today_by_product = [
        {'name': g['_name'], 'color': g['_color'], 'qty': g['today_qty'], 'amount': g['today_amount']}
        for g in group_list if g['today_qty'] > 0
    ]

    top_qty = max(options, key=lambda o: o['today_qty'], default=None)
    top_amount = max(options, key=lambda o: o['today_amount'], default=None)
    # 매출 1등은 오늘 판매가 0이면 None 처리
    if top_qty and top_qty['today_qty'] == 0:
        top_qty = None
    if top_amount and top_amount['today_amount'] == 0:
        top_amount = None

    # 선택일 시간대별(시간 버킷) 매출/판매량 + 시간대별 판매상품 — inventory_log delta(<0) × 가격
    view_start = datetime.combine(sel_date, datetime.min.time())
    view_end = view_start + timedelta(days=1)
    name_color = {m['name']: m['color'] for m in products_meta}
    hour_amount = defaultdict(int)
    hour_qty = defaultdict(int)
    hp_qty = defaultdict(lambda: defaultdict(int))   # hour -> 상품명 -> 수량
    hp_amt = defaultdict(lambda: defaultdict(int))   # hour -> 상품명 -> 금액
    for vid, delta, price, checked_at in (CoupangInventoryLog.objects
            .filter(vendor_item_id__in=vids, checked_at__gte=view_start, checked_at__lt=view_end, delta__lt=0)
            .values_list('vendor_item_id', 'delta', 'sale_price', 'checked_at')):
        sold = -delta
        if sold > ABNORMAL_DROP:    # 비정상 대량감소 제외
            continue
        h = checked_at.hour
        amt = sold * (price or 0)
        hour_qty[h] += sold
        hour_amount[h] += amt
        pn = vid_to_pname.get(vid) or vid
        hp_qty[h][pn] += sold
        hp_amt[h][pn] += amt

    revenue_series = []
    cum = 0
    for h in range(24):
        cum += hour_amount[h]
        revenue_series.append({
            'hour': f'{h:02d}시',
            'amount': hour_amount[h],
            'qty': hour_qty[h],
            'cum_amount': cum,
        })

    # 시간대별 판매상품 리스트 (판매 있는 시간만, 수량 내림차순)
    hourly_products = []
    for h in range(24):
        if not hp_qty[h]:
            continue
        items = sorted(
            ({'name': pn, 'qty': q, 'amount': hp_amt[h][pn], 'color': name_color.get(pn, '#94a3b8')}
             for pn, q in hp_qty[h].items()),
            key=lambda x: -x['qty'])
        hourly_products.append({
            'hour': f'{h:02d}시',
            'total_qty': hour_qty[h],
            'total_amount': hour_amount[h],
            'items': items,
        })

    # 시간대별 판매 패턴 — 선택기간(qty/amount) + 오늘/7일/30일 별도 집계 (비정상 대량감소 제외)
    from datetime import timedelta as _td
    pattern_start_dt = datetime.combine(pattern_start, datetime.min.time())
    today_dt = datetime.combine(today, datetime.min.time())
    d7_dt = datetime.combine(today - _td(days=6), datetime.min.time())
    d30_dt = datetime.combine(today - _td(days=29), datetime.min.time())
    earliest = min(pattern_start_dt, d30_dt)
    pat_qty = defaultdict(int); pat_amount = defaultdict(int)
    today_q = defaultdict(int); week_q = defaultdict(int); month_q = defaultdict(int)
    today_amt = defaultdict(int)
    day_rev = defaultdict(int); day_rev_qty = defaultdict(int)   # 일별 실매출(실제 판매가 기준)
    for delta, price, checked_at in (CoupangInventoryLog.objects
            .filter(vendor_item_id__in=vids, checked_at__gte=earliest, delta__lt=0)
            .values_list('delta', 'sale_price', 'checked_at')):
        sold = -delta
        if sold > ABNORMAL_DROP:    # 비정상 대량감소 제외
            continue
        h = checked_at.hour
        amt = sold * (price or 0)
        if checked_at >= d30_dt:    # 30일 이내만 일별 실매출 집계
            dk = checked_at.date()
            day_rev[dk] += amt
            day_rev_qty[dk] += sold
        if checked_at >= pattern_start_dt:
            pat_qty[h] += sold; pat_amount[h] += amt
        if checked_at >= today_dt:
            today_q[h] += sold; today_amt[h] += amt
        if checked_at >= d7_dt:
            week_q[h] += sold
        if checked_at >= d30_dt:
            month_q[h] += sold

    # 일별 실매출 시리즈 (최근 30일, 실제 판매가 기준)
    daily_revenue = []
    for i in range(30):
        d = month_start + timedelta(days=i)
        daily_revenue.append({
            'date': d.strftime('%m-%d'),
            'full_date': d.strftime('%Y-%m-%d'),
            'amount': day_rev.get(d, 0),
            'qty': day_rev_qty.get(d, 0),
        })
    peak_hour = max(range(24), key=lambda h: pat_qty[h]) if any(pat_qty.values()) else None
    hourly_pattern = [
        {'hour': f'{h:02d}시', 'qty': pat_qty[h], 'amount': pat_amount[h], 'is_peak': h == peak_hour,
         'today_qty': today_q[h], 'today_amount': today_amt[h],
         'week_qty': week_q[h], 'month_qty': month_q[h]}
        for h in range(24)
    ]

    return {
        'today': today.strftime('%Y-%m-%d'),
        'view_date': sel_date.strftime('%Y-%m-%d'),
        'hourly_products': hourly_products,
        'options': options,
        'top_qty': top_qty,
        'top_amount': top_amount,
        'today_total_qty': sum(o['today_qty'] for o in options),
        'today_total_amount': sum(o['today_amount'] for o in options),
        'best_daily': best_daily,
        'best_weekly': best_weekly,
        'best_monthly': best_monthly,
        'products_meta': products_meta,
        'product_daily': product_daily,
        'today_by_product': today_by_product,
        'revenue_series': revenue_series,
        'daily_revenue': daily_revenue,
        'hourly_pattern': hourly_pattern,
        'peak_hour': f'{peak_hour:02d}시' if peak_hour is not None else None,
        'pattern_days': pattern_days,
        # 입고필요 목록(팝업용): 현재고 < 월판매 AND 선입고 미등록
        'restock_needed': [
            {'product_name': o['product_name'], 'option_name': o['option_name'],
             'last_stock': o['last_stock'], 'month_qty': o['month_qty'], 'id': o['id']}
            for o in options if o.get('restock_needed')
        ],
    }


# ── 전 상품 재고 체크 (스트리밍 제너레이터: {"t","m"/"done"}) ──
def check_all_products(account_id: int = None):
    """활성 상품 전체 재고 조회 → NDJSON 이벤트 yield"""
    qs = CoupangRocketProduct.objects.filter(is_active=True).select_related('account')
    if account_id:
        qs = qs.filter(account_id=account_id)
    qs = qs.filter(account__is_active=True)

    total = qs.count()
    yield {"t": "log", "m": f"재고 체크 시작 — 대상 {total}개 옵션"}

    ok, fail = 0, 0
    for i, product in enumerate(qs, 1):
        price, stock, err = get_stock_amount(product.account, product.vendor_item_id)
        if err:
            fail += 1
            yield {"t": "log", "m": f"[{i}/{total}] ❌ {product.vendor_item_id} — {err}"}
        else:
            record_inventory(product, price, stock)
            ok += 1
            yield {"t": "log", "m": f"[{i}/{total}] ✅ {product.vendor_item_id} 재고 {stock} / {price}원 ({product.product_name[:20]})"}
        time.sleep(0.3)  # rate limit 완화

    yield {"t": "log", "m": "일별 판매량/입고 재집계 중(오늘)..."}
    from datetime import date as _date
    rebuild_daily_sales(only_date=_date.today())
    yield {"t": "done", "ok": ok, "fail": fail, "total": total}


# ── 노출상품ID → 형제 옵션 전체 자동등록 (Open API, 브라우저 불필요) ──
SELLER_PRODUCTS_PATH = "/v2/providers/seller_api/apis/api/v1/marketplace/seller-products"
CDN_IMG_BASE = "https://image1.coupangcdn.com/image/"


def _find_seller_product_id(account, displayed_product_id, max_pages=40):
    """노출상품ID(productId) → 등록상품ID(sellerProductId) 매핑.
    목록 API는 createdAt 최신순이라 신규상품은 앞쪽 페이지에서 발견."""
    target = str(displayed_product_id).strip()
    token = ''
    for _ in range(max_pages):
        q = f"vendorId={account.vendor_id}&maxPerPage=50" + (f"&nextToken={token}" if token else "")
        d, err = _signed_get(account, SELLER_PRODUCTS_PATH, q)
        if err:
            return None, f"목록조회 실패: {err}"
        for it in (d.get('data') or []):
            if str(it.get('productId')) == target:
                return it.get('sellerProductId'), None
        token = d.get('nextToken')
        if not token:
            break
    return None, f"목록 {max_pages}페이지 내 노출상품ID {target} 미발견(해당 계정 소유 아님 가능)"


def _rep_image_cdn(item):
    """대표이미지(REPRESENTATION, imageOrder=0) cdnPath 반환."""
    imgs = item.get('images') or []
    for im in imgs:
        if im.get('imageType') == 'REPRESENTATION' and im.get('cdnPath'):
            return im.get('cdnPath')
    for im in imgs:
        if im.get('cdnPath'):
            return im.get('cdnPath')
    return ''


def fetch_product_options(account, displayed_product_id):
    """노출상품ID의 모든 옵션을 Open API 로 조회.
    로켓그로스 옵션ID(rgVid)를 재고추적 기본키로, 판매자윙 옵션ID(mpVid)도 함께 보존.
    반환: (product_name, [{vendor_item_id(=rgVid우선), marketplace_vid(mpVid), rocket_vid(rgVid),
                          option_name, barcode, sku, image_cdn}], error)"""
    spid, err = _find_seller_product_id(account, displayed_product_id)
    if err or not spid:
        return None, [], (err or '등록상품ID 없음')
    d, err = _signed_get(account, f"{SELLER_PRODUCTS_PATH}/{spid}")
    if err:
        return None, [], f"상품상세 실패: {err}"
    data = d.get('data') or {}
    pname = data.get('sellerProductName') or ''
    out = []
    for it in (data.get('items') or []):
        mp = it.get('marketplaceItemData') or {}
        rg = it.get('rocketGrowthItemData') or {}
        mp_vid = str(mp.get('vendorItemId') or '')
        rg_vid = str(rg.get('vendorItemId') or '')
        primary = rg_vid or mp_vid     # 로켓그로스 ID 우선(실재고 추적)
        if not primary:
            continue
        out.append({
            'vendor_item_id': primary,
            'marketplace_vid': mp_vid,
            'rocket_vid': rg_vid,
            'option_name': it.get('itemName') or '',
            'barcode': (rg.get('barcode') or mp.get('barcode') or ''),
            'sku': (rg.get('externalVendorSku') or mp.get('externalVendorSku') or ''),
            'image_cdn': _rep_image_cdn(it),
        })
    return pname, out, None


def verify_and_fix_all_vendor_ids(account=None):
    """등록된 모든 상품의 옵션ID를 검증.
    판매자옵션ID(mpVid)로 잘못 등록된 건 → 로켓상품ID(rgVid)를 자동으로 찾아 교체 +
    판매자옵션ID도 함께 저장. NDJSON 이벤트 yield."""
    from collections import defaultdict
    if account is None:
        account = CoupangApiAccount.objects.filter(
            access_key__gt='', secret_key_enc__gt='').first()
    if not account:
        yield {"t": "log", "m": "❌ Open API 키 보유 계정 없음"}
        yield {"t": "done", "fixed": 0, "ok": 0, "fail": 0, "total": 0}
        return

    bygroup = defaultdict(list)
    for p in CoupangRocketProduct.objects.all():
        bygroup[p.seller_product_id].append(p)

    fixed = ok = fail = skipped = 0
    yield {"t": "log", "m": f"검증 시작 — {CoupangRocketProduct.objects.count()}개 옵션 / {len(bygroup)}개 상품"}

    for spid, prods in bygroup.items():
        if not spid:
            skipped += len(prods)
            yield {"t": "log", "m": f"⏭ [노출ID없음] {prods[0].product_name[:18]} {len(prods)}옵션 — 스킵"}
            continue
        rsp, err = _find_seller_product_id(account, spid)
        if not rsp:
            skipped += len(prods)
            yield {"t": "log", "m": f"⏭ [{spid}] {prods[0].product_name[:18]} — Open API 미발견({err or ''})"}
            continue
        d, e = _signed_get(account, f"{SELLER_PRODUCTS_PATH}/{rsp}")
        if e or not d:
            fail += len(prods)
            yield {"t": "log", "m": f"❌ [{spid}] 상세조회 실패: {e}"}
            continue
        mp2rg = {}; rg_set = set()
        for it in (d.get('data', {}).get('items') or []):
            mp = str((it.get('marketplaceItemData') or {}).get('vendorItemId') or '')
            rg = str((it.get('rocketGrowthItemData') or {}).get('vendorItemId') or '')
            if mp:
                mp2rg[mp] = rg
            if rg:
                rg_set.add(rg)
        for p in prods:
            v = p.vendor_item_id
            label = f"{p.product_name[:16]} · {p.option_name}"
            if v in rg_set:
                # 이미 로켓ID — 판매자옵션ID만 백필
                if not p.marketplace_vendor_item_id:
                    for mpv, rgv in mp2rg.items():
                        if rgv == v:
                            p.marketplace_vendor_item_id = mpv
                            p.save(update_fields=['marketplace_vendor_item_id'])
                            break
                ok += 1
                yield {"t": "log", "m": f"✅ {label} — 로켓ID 정상"}
            elif v in mp2rg and mp2rg[v]:
                # 판매자옵션ID로 잘못 등록 → 안내 메시지 + 로켓ID 자동 전환
                rgv = mp2rg[v]
                yield {"t": "log", "m": f"⚠ 「{label}」 판매자옵션ID 상품입니다. 로켓상품 ID를 자동으로 찾아 둘 다 등록하겠습니다."}
                if CoupangRocketProduct.objects.filter(vendor_item_id=rgv).exclude(id=p.id).exists():
                    fail += 1
                    yield {"t": "log", "m": f"   ↳ ❌ 로켓ID {rgv} 중복 — 스킵"}
                    continue
                CoupangInventoryLog.objects.filter(vendor_item_id=v).delete()
                p.marketplace_vendor_item_id = v
                p.vendor_item_id = rgv
                p.last_stock = p.last_price = p.last_checked_at = None
                p.save(update_fields=['vendor_item_id', 'marketplace_vendor_item_id',
                                      'last_stock', 'last_price', 'last_checked_at'])
                try:
                    pr, st, serr = get_stock_amount(account, rgv)
                    if not serr:
                        record_inventory(p, pr, st)
                        stock_txt = f" 로켓재고 {st}"
                    else:
                        stock_txt = f" (재고조회 {serr[:30]})"
                except Exception:
                    stock_txt = ""
                fixed += 1
                yield {"t": "log", "m": f"   ↳ 🔧 판매자ID {v} → 로켓ID {rgv} 등록완료.{stock_txt}"}
            else:
                fail += 1
                yield {"t": "log", "m": f"❓ {label} — mp/rg 어느쪽도 아님 (vid={v})"}

    yield {"t": "log", "m": f"검증 완료 — 정상 {ok} / 수정 {fixed} / 실패 {fail} / 스킵 {skipped}"}
    yield {"t": "done", "fixed": fixed, "ok": ok, "fail": fail, "total": ok + fixed + fail + skipped}


def register_product_options(account, displayed_product_id, log_fn=None):
    """노출상품ID의 전체 옵션을 CoupangRocketProduct 로 자동등록 + 대표이미지 다운로드.
    이미 등록된 옵션(vendor_item_id 기준)은 정보/이미지만 갱신. NDJSON 친화 반환."""
    import os as _os
    from . import coupang_image_service as _cis

    def log(m):
        if log_fn:
            log_fn(m)

    pname, opts, err = fetch_product_options(account, displayed_product_id)
    if err:
        raise RuntimeError(err)
    log(f"'{pname}' 옵션 {len(opts)}개 조회 완료 — 등록 시작")
    _os.makedirs(_cis.IMG_DIR, exist_ok=True)

    created = updated = img_ok = 0
    results = []
    for o in opts:
        vid = o['vendor_item_id']
        obj, is_new = CoupangRocketProduct.objects.update_or_create(
            vendor_item_id=vid,
            defaults={
                'account': account,
                'marketplace_vendor_item_id': o.get('marketplace_vid', ''),
                'seller_product_id': str(displayed_product_id),
                'product_name': pname,
                'option_name': o['option_name'],
                'barcode': o['barcode'],
            },
        )
        # 대표이미지 CDN 직접 다운로드 (브라우저 불필요)
        img_saved = False
        if o['image_cdn']:
            try:
                url = CDN_IMG_BASE + o['image_cdn']
                idata, ext = _cis._download_image(url)
                fname = f"{vid}.{ext}"
                with open(_os.path.join(_cis.IMG_DIR, fname), 'wb') as f:
                    f.write(idata)
                obj.image_url = url
                obj.image_file = fname
                obj.image_crawled_at = timezone.now()
                obj.save(update_fields=['image_url', 'image_file', 'image_crawled_at'])
                img_ok += 1
                img_saved = True
            except Exception as e:
                log(f"  ⚠ {o['option_name']} 이미지 실패: {str(e)[:50]}")
        # 등록 즉시 1회 재고조회 (대시보드 즉시 반영)
        stock_txt = ''
        try:
            price, stock, serr = get_stock_amount(account, vid)
            if not serr:
                record_inventory(obj, price, stock)
                stock_txt = f' 재고{stock}'
        except Exception:
            pass
        created += 1 if is_new else 0
        updated += 0 if is_new else 1
        tag = '신규' if is_new else '갱신'
        log(f"  {'✅' if is_new else '🔄'} [{tag}] {o['option_name']} (vid={vid}){' +이미지' if img_saved else ''}{stock_txt}")
        results.append({'vendor_item_id': vid, 'option_name': o['option_name'], 'new': is_new, 'image': img_saved})

    log(f"완료 — 신규 {created} / 갱신 {updated} / 이미지 {img_ok}")
    return {'product_name': pname, 'registered': created, 'updated': updated, 'images': img_ok, 'options': results}


# ── 옵션/SKU별 정산 대조 (쿠팡 실제정산 API ↔ 내 대장) ──
def get_coupang_option_settlement(cupang_id, date_from, date_to):
    """쿠팡 revenue-history(실제정산, 옵션ID별) ↔ order DB(대장, S코드별) 를
    SKU(=S코드) 기준으로 묶어 대조. 쿠팡 API 키가 있는 계정만 가능(현재 exansys).
    반환: {rows, totals, error}. rows = SKU별 {api_*, odb_*, diff, vids[]}"""
    from collections import defaultdict
    acc = CoupangApiAccount.objects.filter(cupang_id=cupang_id).first()
    if not acc:
        return {'error': f'계정 없음: {cupang_id}', 'rows': [], 'totals': {}}
    if not (acc.access_key and acc.secret_key_enc):
        return {'error': f'{cupang_id}: 쿠팡 Open API 키 미등록(실제정산 조회 불가)', 'rows': [], 'totals': {}}

    # 1) 쿠팡 실제정산 — 옵션ID(vendorItemId)별
    rev = get_revenue_history(acc, date_from, date_to)
    if rev.get('error') and not rev['rows']:
        return {'error': rev['error'], 'rows': [], 'totals': {}}

    # 매칭키 = 노출상품ID(productId). 옵션ID/SKU는 반품·로켓·윙마다 새로 생기지만
    # 노출상품ID는 고정 → 노출상품ID 기준으로 묶어야 정확히 대조됨.
    # productId(API) == 대장 product_code == CoupangRocketProduct.seller_product_id
    by_exp = defaultdict(lambda: {'name': '', 'qty': 0, 'sale': 0, 'fee': 0, 'settle': 0, 'vids': {}, 'skus': set()})
    for x in rev['rows']:
        vid = x['vendor_item_id']
        sku = (x['sku'] or '').strip()
        exp = str(x.get('product_id') or '').strip()
        key = exp or (sku and f'SKU:{sku}') or f'(노출없음:{vid})'
        s = by_exp[key]
        s['name'] = x['product_name'] or s['name']
        s['qty'] += x['quantity']; s['sale'] += x['sale_amount']
        s['fee'] += x['service_fee']; s['settle'] += x['settlement_amount']
        if sku:
            s['skus'].add(sku)
        if vid:
            v = s['vids'].setdefault(vid, {'qty': 0, 'settle': 0})
            v['qty'] += x['quantity']; v['settle'] += x['settlement_amount']

    # 2) 내 대장 — 노출상품ID(product_code)별 (반품·취소 제외)
    #    product_code가 비어있는 행은 같은 S코드(product_seller_code)의 노출상품ID로 보정
    RET = ("(order_status LIKE '%%반품%%' OR order_status LIKE '%%취소%%' "
           "OR order_status LIKE '%%환불%%' OR order_status LIKE '%%교환%%')")
    alias = _cupang_to_alias(cupang_id)
    odb = defaultdict(lambda: {'qty': 0, 'settle': 0, 'pay': 0, 'name': '', 'codes': set(),
                               'hours': defaultdict(int)})  # hours: 구매시간대(order_time) 수량분포
    with connections['joacham'].cursor() as cur:
        where = ["site_name='06.쿠팡'", "order_date>=%s", "order_date<=%s"]
        params = [date_from, date_to + ' 23:59:59']
        if alias:
            where.append("seller_alias=%s"); params.append(alias)
        cur.execute(f"""
            SELECT product_seller_code, product_code,
                   SUM(CASE WHEN NOT {RET} THEN quantity ELSE 0 END),
                   SUM(CASE WHEN NOT {RET} THEN settlement_price ELSE 0 END),
                   SUM(CASE WHEN NOT {RET} THEN payment_price ELSE 0 END),
                   MAX(product_name)
            FROM orders_order WHERE {' AND '.join(where)}
            GROUP BY product_seller_code, product_code""", params)
        ledger_rows = [(str(r[0] or '').strip(), str(r[1] or '').strip(),
                        int(r[2] or 0), int(r[3] or 0), int(r[4] or 0), r[5] or '')
                       for r in cur.fetchall()]
        # 구매시간대 — 판매자배송 주문 시각(order_time) 기준 시간대별 구매수량
        cur.execute(f"""
            SELECT product_seller_code, product_code, HOUR(order_time) AS h,
                   SUM(CASE WHEN NOT {RET} THEN quantity ELSE 0 END)
            FROM orders_order WHERE {' AND '.join(where)} AND order_time IS NOT NULL
            GROUP BY product_seller_code, product_code, h""", params)
        hour_rows = [(str(r[0] or '').strip(), str(r[1] or '').strip(),
                      int(r[2]) if r[2] is not None else None, int(r[3] or 0)) for r in cur.fetchall()]
        # S코드 → 노출상품ID 자동학습 (전체 이력, 기간 무관) — 기간 내 product_code가 비어도 보정되도록
        where2 = ["site_name='06.쿠팡'", "product_seller_code<>''", "product_code<>''"]
        params2 = []
        if alias:
            where2.append("seller_alias=%s"); params2.append(alias)
        cur.execute(f"""
            SELECT product_seller_code, product_code, COUNT(*) c
            FROM orders_order WHERE {' AND '.join(where2)}
            GROUP BY product_seller_code, product_code ORDER BY c DESC""", params2)
        code2exp = {}
        for sc, pc, _c in cur.fetchall():
            sc = str(sc).strip(); pc = str(pc).strip()
            if sc and pc and sc not in code2exp:   # 가장 많이 쓰인 노출ID 채택(ORDER BY c DESC)
                code2exp[sc] = pc
    def _resolve_exp(sc, pc):
        return pc or code2exp.get(sc) or (sc and f'SKU:{sc}') or '(미분류)'
    for sc, pc, q, s, pay, nm in ledger_rows:
        d = odb[_resolve_exp(sc, pc)]
        d['qty'] += q; d['settle'] += s; d['pay'] += pay
        d['name'] = d['name'] or nm
        if sc:
            d['codes'].add(sc)
    for sc, pc, h, q in hour_rows:
        if h is None or q <= 0:
            continue
        odb[_resolve_exp(sc, pc)]['hours'][h] += q

    # 3) 합치기 (노출상품ID 기준 합집합)
    keys = set(by_exp) | set(odb)
    rows = []
    for k in keys:
        a = by_exp.get(k)
        o = odb.get(k)
        api_settle = a['settle'] if a else 0
        odb_settle = o['settle'] if o else 0
        if a and o:
            match = 'both'
        elif a:
            match = 'api_only'   # 쿠팡 정산엔 있는데 대장에 없음
        else:
            match = 'odb_only'   # 대장엔 있는데 쿠팡 정산 인식 안됨(시점차 등)
        # 대표 S코드: API SKU 우선, 없으면 대장 S코드
        skus = sorted((a['skus'] if a else set()) | (o['codes'] if o else set()))
        sku_label = ', '.join(skus) if skus else '(SKU미등록)'
        # 판매자배송(주문DB) 구매시간대 — 시간대별 구매수량 + 피크시간
        hours = o['hours'] if o else {}
        hour_list = [{'hour': h, 'qty': hours[h]} for h in sorted(hours)]
        peak_hour = max(hours, key=hours.get) if hours else None
        rows.append({
            'sku': sku_label,
            'exposure_id': k if not k.startswith(('SKU:', '(')) else '',
            'product_name': (a['name'] if a else '') or (o['name'] if o else ''),
            'vids': sorted((a['vids'].keys() if a else []), key=str),
            'api_qty': a['qty'] if a else 0,
            'api_sale': a['sale'] if a else 0,
            'api_fee': a['fee'] if a else 0,
            'api_settle': api_settle,
            'odb_qty': o['qty'] if o else 0,
            'odb_settle': odb_settle,
            'odb_pay': o['pay'] if o else 0,          # 판매자배송 구매금액(결제가 합)
            'peak_hour': peak_hour,                    # 구매 피크 시간대
            'hours': hour_list,                        # 시간대별 구매수량 분포
            'diff': api_settle - odb_settle,
            'match': match,
        })
    rows.sort(key=lambda r: (-(r['api_settle'] or r['odb_settle'])))

    tot = {
        'sku_count': len(rows),
        'api_qty': sum(r['api_qty'] for r in rows),
        'api_sale': sum(r['api_sale'] for r in rows),
        'api_fee': sum(r['api_fee'] for r in rows),
        'api_settle': sum(r['api_settle'] for r in rows),
        'odb_qty': sum(r['odb_qty'] for r in rows),
        'odb_settle': sum(r['odb_settle'] for r in rows),
        'odb_pay': sum(r['odb_pay'] for r in rows),
        'diff': sum(r['diff'] for r in rows),
        'both': sum(1 for r in rows if r['match'] == 'both'),
        'api_only': sum(1 for r in rows if r['match'] == 'api_only'),
        'odb_only': sum(1 for r in rows if r['match'] == 'odb_only'),
    }
    return {'rows': rows, 'totals': tot}


def _cupang_to_alias(cupang_id):
    """쿠팡 로그인ID → order DB seller_alias 추정 (06.쿠팡 대장에서 가장 많이 쓰인 alias)."""
    # exansys → 13엑사엔시스 등. 대장에서 직접 못 찾으면 None(전 사업자)
    _MAP = {'exansys': '13엑사엔시스', 'bitcom1': '03비트컴', 'bitic05': '05비트윙',
            'joacham': '14조아참', 'nainjoy6': '06나인조이'}
    return _MAP.get(cupang_id)
